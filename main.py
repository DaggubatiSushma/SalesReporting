from __future__ import annotations

import argparse
import calendar
import os
import secrets
import sqlite3
import threading
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from flask import Flask, abort, g, jsonify, request, send_file, send_from_directory, session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter
from werkzeug.exceptions import HTTPException
from dotenv import load_dotenv

try:
    import psycopg
except ImportError:  # pragma: no cover
    psycopg = None


BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR / "frontend"
DATABASE_PATH = BASE_DIR / "sales_reporting.db"
load_dotenv(BASE_DIR / ".env")
DATABASE_URL = os.environ.get("DATABASE_URL")
USE_POSTGRES = bool(DATABASE_URL)
SCHEMA_PATH = (
    (BASE_DIR / "backend" / ("schema.postgres.sql" if USE_POSTGRES else "schema.sql")).resolve()
    if (BASE_DIR / "backend").exists()
    else (BASE_DIR / ("schema.postgres.sql" if USE_POSTGRES else "schema.sql")).resolve()
)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "sales-reporting-secret-change-me")
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

MASTER_USERNAME = os.environ.get("MASTER_USERNAME", "master")
MASTER_PASSWORD = os.environ.get("MASTER_PASSWORD", "Master@123")

DAILY_SALES_MAPPING = {
    "date": "entry_date",
    "total": "total",
    "sales": "sales",
    "hst": "hst",
    "online": "online",
    "instant": "instant",
    "cc": "cc",
    "gc": "gc",
    "nonAdd": "non_add",
    "mc": "mc",
    "visa": "visa",
    "debit": "debit",
    "cash": "cash",
    "lotteryPayment": "lottery_payment",
    "lotteryIncome": "lottery_income",
}

RESOURCE_CONFIG = {
    "lottery-records": {
        "table": "lottery_records",
        "date_col": "lottery_date",
        "mapping": {
            "date": "lottery_date",
            "lotteryPayment": "lottery_payment",
            "lotteryIncome": "lottery_income",
        },
        "required": ("date",),
        "numeric": {"lotteryPayment", "lotteryIncome"},
    },
    "cash-payments": {
        "table": "cash_payments",
        "date_col": "payment_date",
        "mapping": {
            "date": "payment_date",
            "vendorName": "vendor_name",
            "amount": "amount",
            "hst": "hst",
        },
        "required": ("date", "vendorName"),
        "numeric": {"amount", "hst"},
    },
    "bank-payments": {
        "table": "bank_payments",
        "date_col": "payment_date",
        "mapping": {
            "date": "payment_date",
            "vendorName": "vendor_name",
            "amount": "amount",
            "hst": "hst",
            "chq": "chq",
        },
        "required": ("date", "vendorName"),
        "numeric": {"amount", "hst"},
    },
    "expenses": {
        "table": "expenses",
        "date_col": "expense_date",
        "mapping": {
            "date": "expense_date",
            "vendorName": "vendor_name",
            "amount": "amount",
            "hst": "hst",
        },
        "required": ("date", "vendorName"),
        "numeric": {"amount", "hst"},
    },
    "salaries": {
        "table": "salaries",
        "date_col": "salary_date",
        "mapping": {
            "date": "salary_date",
            "employee": "employee",
            "amount": "amount",
        },
        "required": ("date", "employee"),
        "numeric": {"amount"},
    },
    "other-income": {
        "table": "other_income",
        "date_col": "income_date",
        "mapping": {
            "date": "income_date",
            "vendorName": "vendor_name",
            "amount": "amount",
        },
        "required": ("date", "vendorName"),
        "numeric": {"amount"},
    },
    "lcbo-entries": {
        "table": "lcbo_entries",
        "date_col": "entry_date",
        "mapping": {
            "date": "entry_date",
            "vendorName": "vendor_name",
            "invoiceNo": "invoice_no",
            "creditEnding": "credit_ending",
            "amount": "amount",
            "hst": "hst",
        },
        "required": ("date", "vendorName", "invoiceNo"),
        "numeric": {"amount", "hst"},
    },
    "credit-card-payments": {
        "table": "credit_card_payments",
        "date_col": "payment_date",
        "mapping": {
            "date": "payment_date",
            "purpose": "purpose",
            "amount": "amount",
        },
        "required": ("date", "purpose"),
        "numeric": {"amount"},
    },
}

MONTH_NAMES = (
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
)

DATASET_DATE_FIELDS = {
    "sales_entries": "entry_date",
    "lottery_records": "lottery_date",
    "cash_payments": "payment_date",
    "bank_payments": "payment_date",
    "expenses": "expense_date",
    "salaries": "salary_date",
    "other_income": "income_date",
    "lcbo_entries": "entry_date",
    "credit_card_payments": "payment_date",
}

EXCEL_NUMBER_FORMAT = "#,##0.00;-#,##0.00;0"
EXCEL_DATE_FORMAT = "yyyy-mm-dd"
EXCEL_PERCENT_FORMAT = "0.00"
CCR_SOURCE_TYPE = "credit-card-reconciliation"
CCR_STATUS_UNALLOCATED = "UNALLOCATED"
CCR_STATUS_ALLOCATED = "ALLOCATED"
CCR_STATUS_REVERSED = "REVERSED"
LCBO_SOURCE_TYPE = "lcbo-monthly"
LCBO_STATUS_PENDING = "PENDING_VALIDATION"
LCBO_STATUS_VALIDATED = "VALIDATED"
LCBO_STATUS_POSTED = "POSTED_TO_CASH_DEBIT"
LCBO_LOCKED_RESOURCES = {"lcbo-entries", "credit-card-payments"}
DEFAULT_STORE_NAMES = ["Llyod", "Dundas", "Lakeshore", "Mimico", "Skymark", "GG", "Kingsbury", "Albion"]
STORE_BUTTON_LIMIT = len(DEFAULT_STORE_NAMES)
_db_initialized = False
_db_init_lock = threading.Lock()
EXPORT_MONTH_CODES = tuple(month[:3].upper() for month in MONTH_NAMES)
EXPORT_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
EXPORT_GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
EXPORT_HEADER_FILL = PatternFill("solid", fgColor="B4C6E7")
EXPORT_TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
EXPORT_HIGHLIGHT_FILL = PatternFill("solid", fgColor="BDD7EE")
EXPORT_BORDER = Border(
    left=Side(style="thin", color="94A3B8"),
    right=Side(style="thin", color="94A3B8"),
    top=Side(style="thin", color="94A3B8"),
    bottom=Side(style="thin", color="94A3B8"),
)


class PgResultProxy:
    def __init__(self, cursor):
        self._cursor = cursor

    @property
    def rowcount(self):
        return self._cursor.rowcount

    @property
    def lastrowid(self):
        return self._cursor.lastrowid

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchall(self):
        return self._cursor.fetchall()


class PgConnectionProxy:
    def __init__(self, connection):
        self._connection = connection

    def execute(self, query, params=()):
        normalized = query.replace("?", "%s")
        cursor = self._connection.cursor(row_factory=psycopg.rows.dict_row)
        cursor.execute(normalized, params)
        return PgResultProxy(cursor)

    def commit(self):
        self._connection.commit()

    def rollback(self):
        self._connection.rollback()

    def close(self):
        self._connection.close()


def run_sql_script(connection, sql_script: str) -> None:
    statements: list[str] = []
    current: list[str] = []
    quote_char: str | None = None
    escape = False
    for char in sql_script:
        if escape:
            current.append(char)
            escape = False
            continue
        if quote_char:
            current.append(char)
            if char == "\\":
                escape = True
            elif char == quote_char:
                quote_char = None
            continue
        if char in {"'", '"'}:
            quote_char = char
            current.append(char)
            continue
        if char == ";":
            statement = "".join(current).strip()
            if statement:
                statements.append(statement)
            current = []
            continue
        current.append(char)
    tail = "".join(current).strip()
    if tail:
        statements.append(tail)
    with connection.cursor() as cursor:
        for statement in statements:
            if not statement.strip():
                continue
            cursor.execute(statement)


def get_db():
    ensure_app_initialized()
    if "db" not in g:
        if USE_POSTGRES:
            if psycopg is None:
                raise RuntimeError("psycopg is required when DATABASE_URL is configured.")
            connection = psycopg.connect(DATABASE_URL, sslmode="require")
            connection.autocommit = False
            g.db = PgConnectionProxy(connection)
        else:
            connection = sqlite3.connect(DATABASE_PATH, timeout=30)
            connection.row_factory = sqlite3.Row
            connection.execute("PRAGMA busy_timeout = 30000")
            connection.execute("PRAGMA foreign_keys = ON")
            g.db = connection
    return g.db


@app.teardown_appcontext
def close_db(_: BaseException | None) -> None:
    connection = g.pop("db", None)
    if connection is not None:
        connection.close()


def init_db() -> None:
    def ensure_table_column(connection, table: str, column: str, definition: str) -> None:
        if USE_POSTGRES:
            columns = {
                row["column_name"]
                for row in connection.execute(
                    """
                    SELECT column_name
                      FROM information_schema.columns
                     WHERE table_schema = current_schema()
                       AND table_name = %s
                    """,
                    (table,),
                ).fetchall()
            }
            if column not in columns:
                connection.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
            return

        columns = {row[1] for row in connection.execute(f"PRAGMA table_info({table})").fetchall()}
        if column not in columns:
            connection.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")

    if USE_POSTGRES:
        connection = psycopg.connect(DATABASE_URL, sslmode="require")
        connection.autocommit = False
        run_sql_script(connection, SCHEMA_PATH.read_text(encoding="utf-8"))
        ensure_table_column(connection, "cash_payments", "source_type", "TEXT NOT NULL DEFAULT ''")
        ensure_table_column(connection, "cash_payments", "source_ref_id", "INTEGER")
        ensure_table_column(connection, "cash_payments", "source_note", "TEXT NOT NULL DEFAULT ''")
        ensure_table_column(connection, "bank_payments", "source_type", "TEXT NOT NULL DEFAULT ''")
        ensure_table_column(connection, "bank_payments", "source_ref_id", "INTEGER")
        ensure_table_column(connection, "bank_payments", "source_note", "TEXT NOT NULL DEFAULT ''")
        ensure_table_column(connection, "credit_card_reconciliation", "hst_cents", "INTEGER NOT NULL DEFAULT 0")
        ensure_table_column(connection, "lcbo_monthly_workflows", "validated_amount", "REAL NOT NULL DEFAULT 0")
        ensure_table_column(connection, "lcbo_monthly_workflows", "validated_at", "TEXT")
        ensure_table_column(connection, "lcbo_monthly_workflows", "notes", "TEXT NOT NULL DEFAULT ''")
        ensure_table_column(connection, "lcbo_monthly_workflows", "posted_at", "TEXT")
        ensure_table_column(connection, "lcbo_monthly_workflows", "posted_resource", "TEXT")
        ensure_table_column(connection, "lcbo_monthly_workflows", "posted_record_id", "INTEGER")
        ensure_table_column(connection, "lcbo_monthly_workflows", "posted_payment_type", "TEXT")
        connection.commit()
        connection.close()
        return

    DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DATABASE_PATH, timeout=30)
    connection.execute("PRAGMA busy_timeout = 30000")
    connection.execute("PRAGMA foreign_keys = ON")
    connection.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
    ensure_table_column(connection, "cash_payments", "source_type", "TEXT NOT NULL DEFAULT ''")
    ensure_table_column(connection, "cash_payments", "source_ref_id", "INTEGER")
    ensure_table_column(connection, "cash_payments", "source_note", "TEXT NOT NULL DEFAULT ''")
    ensure_table_column(connection, "bank_payments", "source_type", "TEXT NOT NULL DEFAULT ''")
    ensure_table_column(connection, "bank_payments", "source_ref_id", "INTEGER")
    ensure_table_column(connection, "bank_payments", "source_note", "TEXT NOT NULL DEFAULT ''")
    ensure_table_column(connection, "credit_card_reconciliation", "hst_cents", "INTEGER NOT NULL DEFAULT 0")
    ensure_table_column(connection, "lcbo_monthly_workflows", "validated_amount", "REAL NOT NULL DEFAULT 0")
    ensure_table_column(connection, "lcbo_monthly_workflows", "validated_at", "TEXT")
    ensure_table_column(connection, "lcbo_monthly_workflows", "notes", "TEXT NOT NULL DEFAULT ''")
    ensure_table_column(connection, "lcbo_monthly_workflows", "posted_at", "TEXT")
    ensure_table_column(connection, "lcbo_monthly_workflows", "posted_resource", "TEXT")
    ensure_table_column(connection, "lcbo_monthly_workflows", "posted_record_id", "INTEGER")
    ensure_table_column(connection, "lcbo_monthly_workflows", "posted_payment_type", "TEXT")

    store_count = connection.execute("SELECT COUNT(*) FROM stores").fetchone()[0]
    if store_count == 0:
        for store_name in DEFAULT_STORE_NAMES:
            connection.execute("INSERT INTO stores (name) VALUES (?)", (store_name,))
    elif store_count < STORE_BUTTON_LIMIT:
        for store_name in DEFAULT_STORE_NAMES:
            exists = connection.execute("SELECT 1 FROM stores WHERE name = ? LIMIT 1", (store_name,)).fetchone()
            if exists is None:
                connection.execute("INSERT INTO stores (name) VALUES (?)", (store_name,))

    connection.commit()
    connection.close()


def ensure_app_initialized() -> None:
    global _db_initialized
    if _db_initialized:
        return
    with _db_init_lock:
        if _db_initialized:
            return
        init_db()
        _db_initialized = True


def api_error(message: str, status: int = 400):
    return jsonify({"error": message}), status


@app.errorhandler(Exception)
def handle_exception(error: Exception):
    if request.path.startswith("/api/"):
        if isinstance(error, HTTPException):
            return api_error(error.description, error.code or 500)
        return api_error(str(error), 500)
    raise error


@app.after_request
def apply_no_cache_headers(response):
    if request.path.startswith("/api/") or request.path == "/":
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


def is_authenticated() -> bool:
    return bool(session.get("authenticated"))


def get_auth_key() -> str | None:
    if not is_authenticated():
        return None
    value = session.get("auth_key")
    return value if isinstance(value, str) and value else None


def is_public_api_path(path: str) -> bool:
    return path in {"/api/auth/status", "/api/auth/login", "/api/auth/logout"}


@app.before_request
def require_api_authentication():
    if request.path.startswith("/api/") and not is_public_api_path(request.path):
        if not is_authenticated():
            abort(401, description="Authentication required")


def normalize_text(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""


def parse_amount(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def parse_money_to_cents(value: Any) -> int:
    if value in (None, ""):
        abort(400, description="Amount is required")
    try:
        amount = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        abort(400, description="Invalid amount")
    if amount <= 0:
        abort(400, description="Amount must be greater than zero")
    return int((amount * Decimal("100")).to_integral_value(rounding=ROUND_HALF_UP))


def parse_non_negative_money_to_cents(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        amount = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        abort(400, description="Invalid HST amount")
    if amount < 0:
        abort(400, description="HST amount cannot be negative")
    return int((amount * Decimal("100")).to_integral_value(rounding=ROUND_HALF_UP))


def cents_to_amount(value: Any) -> float:
    cents = int(value or 0)
    return float((Decimal(cents) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def validate_iso_date(value: str, field_label: str) -> str:
    text = normalize_text(value)
    if not text:
        abort(400, description=f"{field_label} is required")
    try:
        datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        abort(400, description=f"{field_label} must use YYYY-MM-DD format")
    return text


def validate_sales_totals(payload: dict[str, Any]) -> None:
    total = parse_amount(payload.get("total"))
    sales = parse_amount(payload.get("sales"))
    hst = parse_amount(payload.get("hst"))
    if abs(total - (sales + hst)) > 0.009:
        abort(400, description="Total must equal Sales + HST")


def require_fields(payload: dict[str, Any], fields: tuple[str, ...]) -> None:
    missing = [field for field in fields if normalize_text(payload.get(field)) == ""]
    if missing:
        abort(400, description=f"Missing required fields: {', '.join(missing)}")


def month_filter_sql(date_col: str) -> str:
    if USE_POSTGRES:
        return f"substring({date_col}::text from 1 for 7)"
    return f"substr({date_col}, 1, 7)"


def year_filter_sql(date_col: str) -> str:
    if USE_POSTGRES:
        return f"EXTRACT(YEAR FROM {date_col}::date)"
    return f"substr({date_col}, 1, 4)"


def month_numeric_sql(date_col: str) -> str:
    if USE_POSTGRES:
        return f"CAST(EXTRACT(MONTH FROM {date_col}::date) AS INTEGER)"
    return f"CAST(substr({date_col}, 6, 2) AS INTEGER)"


def build_period_filters(
    date_col: str,
    year: int | None = None,
    month: int | None = None,
    quarter: int | None = None,
) -> tuple[list[str], list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []
    if year and month:
        clauses.append(f"{month_filter_sql(date_col)} = ?")
        params.append(f"{year:04d}-{month:02d}")
    elif year and quarter:
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        clauses.append(f"{year_filter_sql(date_col)} = ?")
        clauses.append(f"{month_numeric_sql(date_col)} BETWEEN ? AND ?")
        params.extend([f"{year:04d}", start_month, end_month])
    elif year:
        clauses.append(f"{year_filter_sql(date_col)} = ?")
        params.append(f"{year:04d}")

    return clauses, params


def build_date_filters(store_id: int, date_col: str) -> tuple[str, list[Any]]:
    clauses = ["store_id = ?"]
    params: list[Any] = [store_id]
    period_clauses, period_params = build_period_filters(
        date_col,
        request.args.get("year", type=int),
        request.args.get("month", type=int),
        request.args.get("quarter", type=int),
    )
    clauses.extend(period_clauses)
    params.extend(period_params)
    return " AND ".join(clauses), params


def serialize_daily_sales(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "date": row["entry_date"],
        "total": row["total"],
        "sales": row["sales"],
        "hst": row["hst"],
        "online": row["online"],
        "instant": row["instant"],
        "cc": row["cc"],
        "gc": row["gc"],
        "nonAdd": row["non_add"],
        "mc": row["mc"],
        "visa": row["visa"],
        "debit": row["debit"],
        "cash": row["cash"],
        "lotteryPayment": row["lottery_payment"],
        "lotteryIncome": row["lottery_income"],
    }


def serialize_resource(resource: str, row: sqlite3.Row) -> dict[str, Any]:
    mapping = RESOURCE_CONFIG[resource]["mapping"]
    inverse = {db_key: api_key for api_key, db_key in mapping.items()}
    data = {"id": row["id"]}
    for db_key, api_key in inverse.items():
        data[api_key] = row[db_key]
    return data


def parse_optional_store_id(value: Any, field_label: str) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        store_id = int(value)
    else:
        text = normalize_text(value)
        if text == "":
            return None
        try:
            store_id = int(text)
        except ValueError:
            abort(400, description=f"{field_label} must be a valid store id")
    ensure_store_exists(store_id)
    return store_id


def serialize_credit_card_reconciliation(row: sqlite3.Row) -> dict[str, Any]:
    amount = cents_to_amount(row["amount_cents"])
    hst = cents_to_amount(row["hst_cents"])
    return {
        "id": row["id"],
        "transactionDate": row["transaction_date"],
        "creditCard": row["credit_card"],
        "merchant": row["merchant_name"],
        "description": row["description"],
        "amount": amount,
        "amountCents": row["amount_cents"],
        "hst": hst,
        "hstCents": row["hst_cents"],
        "total": round(amount + hst, 2),
        "totalCents": row["amount_cents"] + row["hst_cents"],
        "dedicatedStoreId": row["dedicated_store_id"],
        "dedicatedStoreName": row["dedicated_store_name"],
        "allocatedStoreId": row["allocated_store_id"],
        "allocatedStoreName": row["allocated_store_name"],
        "status": row["status"],
        "paymentType": row["payment_type"],
        "allocationResource": row["allocation_resource"],
        "allocationRecordId": row["allocation_record_id"],
        "allocatedAt": row["allocated_at"],
        "reversedAt": row["reversed_at"],
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }


def get_credit_card_reconciliation_row(record_id: int) -> sqlite3.Row:
    row = get_db().execute(
        """
        SELECT c.id, c.transaction_date, c.credit_card, c.merchant_name, c.description, c.amount_cents, c.hst_cents,
               c.dedicated_store_id, c.status, c.allocated_store_id, c.payment_type, c.allocation_resource,
               c.allocation_record_id, c.allocated_at, c.reversed_at, c.created_at, c.updated_at,
               ds.name AS dedicated_store_name, alloc.name AS allocated_store_name
          FROM credit_card_reconciliation c
          LEFT JOIN stores ds ON ds.id = c.dedicated_store_id
          LEFT JOIN stores alloc ON alloc.id = c.allocated_store_id
         WHERE c.id = ?
        """,
        (record_id,),
    ).fetchone()
    if row is None:
        abort(404, description="Credit card reconciliation transaction not found")
    return row


def map_payload(payload: dict[str, Any], mapping: dict[str, str], numeric_fields: set[str], partial: bool) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for api_field, db_field in mapping.items():
        if partial and api_field not in payload:
            continue
        value = payload.get(api_field)
        if api_field in numeric_fields:
            record[db_field] = parse_amount(value)
        else:
            record[db_field] = normalize_text(value)
    return record


def get_store_row(store_id: int) -> sqlite3.Row:
    row = get_db().execute("SELECT id, name FROM stores WHERE id = ?", (store_id,)).fetchone()
    if row is None:
        abort(404, description="Store not found")
    return row


def ensure_store_exists(store_id: int) -> None:
    get_store_row(store_id)


def get_resource_config(resource: str) -> dict[str, Any]:
    config = RESOURCE_CONFIG.get(resource)
    if config is None:
        abort(404, description="Resource not found")
    return config


def round_money(value: Any) -> float:
    try:
        amount = Decimal(str(value if value not in (None, "") else 0))
    except InvalidOperation:
        amount = Decimal("0")
    return float(amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def parse_non_negative_amount(value: Any, field_label: str, *, required: bool = False) -> float:
    if value in (None, ""):
        if required:
            abort(400, description=f"{field_label} is required")
        return 0.0
    try:
        amount = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        abort(400, description=f"Invalid {field_label.lower()}")
    if amount < 0:
        abort(400, description=f"{field_label} cannot be negative")
    return float(amount)


def parse_required_int(value: Any, field_label: str) -> int:
    if value in (None, ""):
        abort(400, description=f"{field_label} is required")
    try:
        return int(value)
    except (TypeError, ValueError):
        abort(400, description=f"{field_label} must be a valid number")


def validate_year_month(year: int | None, month: int | None) -> tuple[int, int]:
    if year is None:
        abort(400, description="Year is required")
    if month is None:
        abort(400, description="Month is required")
    if month < 1 or month > 12:
        abort(400, description="Month must be between 1 and 12")
    if year < 2000 or year > 2100:
        abort(400, description="Year must be between 2000 and 2100")
    return year, month


def parse_year_month_from_date(date_value: str, field_label: str = "Date") -> tuple[int, int]:
    validated = validate_iso_date(date_value, field_label)
    return int(validated[0:4]), int(validated[5:7])


def month_period_key(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def month_label(year: int, month: int) -> str:
    return f"{MONTH_NAMES[month - 1]} {year}"


def lcbo_status_label(status: str) -> str:
    if status == LCBO_STATUS_VALIDATED:
        return "Validated"
    if status == LCBO_STATUS_POSTED:
        return "Posted to Cash/Debit"
    return "Pending Validation"


def get_lcbo_workflow_row(db: sqlite3.Connection, store_id: int, year: int, month: int) -> sqlite3.Row | None:
    return db.execute(
        """
        SELECT id, store_id, year, month, status, validated_amount, validated_at, notes,
               posted_at, posted_resource, posted_record_id, posted_payment_type, created_at, updated_at
          FROM lcbo_monthly_workflows
         WHERE store_id = ? AND year = ? AND month = ?
        """,
        (store_id, year, month),
    ).fetchone()


def ensure_lcbo_workflow_row(db: sqlite3.Connection, store_id: int, year: int, month: int) -> sqlite3.Row:
    row = get_lcbo_workflow_row(db, store_id, year, month)
    if row is not None:
        return row
    cursor = db.execute(
        """
        INSERT INTO lcbo_monthly_workflows (store_id, year, month, status, validated_amount)
        VALUES (?, ?, ?, ?, 0)
        """,
        (store_id, year, month, LCBO_STATUS_PENDING),
    )
    return db.execute(
        """
        SELECT id, store_id, year, month, status, validated_amount, validated_at, notes,
               posted_at, posted_resource, posted_record_id, posted_payment_type, created_at, updated_at
          FROM lcbo_monthly_workflows
         WHERE id = ?
        """,
        (cursor.lastrowid,),
    ).fetchone()


def record_lcbo_workflow_event(
    db: sqlite3.Connection,
    workflow_id: int,
    event_type: str,
    from_status: str | None,
    to_status: str | None,
    event_amount: float,
    event_note: str = "",
) -> None:
    db.execute(
        """
        INSERT INTO lcbo_monthly_workflow_events (
            workflow_id, event_type, from_status, to_status, event_amount, event_note
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        (workflow_id, event_type, from_status, to_status, round_money(event_amount), normalize_text(event_note)),
    )


def reset_lcbo_workflow_to_pending_if_validated(
    db: sqlite3.Connection,
    store_id: int,
    year: int,
    month: int,
    reason: str,
) -> None:
    workflow = get_lcbo_workflow_row(db, store_id, year, month)
    if workflow is None:
        return
    if workflow["status"] == LCBO_STATUS_POSTED:
        abort(
            409,
            description=(
                f"LCBO month {month_label(year, month)} is already posted. "
                "Reverse the posted LCBO payment before changing LCBO records."
            ),
        )
    if workflow["status"] == LCBO_STATUS_VALIDATED:
        db.execute(
            """
            UPDATE lcbo_monthly_workflows
               SET status = ?,
                   validated_amount = 0,
                   validated_at = NULL,
                   notes = '',
                   updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (LCBO_STATUS_PENDING, workflow["id"]),
        )
        record_lcbo_workflow_event(
            db,
            workflow["id"],
            "DATA_CHANGED",
            LCBO_STATUS_VALIDATED,
            LCBO_STATUS_PENDING,
            0,
            reason,
        )


def enforce_lcbo_month_editable(db: sqlite3.Connection, store_id: int, date_value: str, reason: str) -> None:
    year, month = parse_year_month_from_date(date_value)
    reset_lcbo_workflow_to_pending_if_validated(db, store_id, year, month, reason)


def fetch_lcbo_month_payload(store_id: int, year: int, month: int) -> dict[str, Any]:
    db = get_db()
    store = dict(get_store_row(store_id))
    period_key = month_period_key(year, month)

    lcbo_rows = db.execute(
        f"""
        SELECT id, entry_date, vendor_name, invoice_no, credit_ending, amount, hst
          FROM lcbo_entries
         WHERE store_id = ? AND {month_filter_sql('entry_date')} = ?
         ORDER BY entry_date ASC, id ASC
        """,
        (store_id, period_key),
    ).fetchall()
    payment_rows = db.execute(
        f"""
        SELECT id, payment_date, purpose, amount
          FROM credit_card_payments
         WHERE store_id = ? AND {month_filter_sql('payment_date')} = ?
         ORDER BY payment_date ASC, id ASC
        """,
        (store_id, period_key),
    ).fetchall()

    lcbo_records = [
        {
            "id": row["id"],
            "date": row["entry_date"],
            "vendorName": row["vendor_name"],
            "invoiceNo": row["invoice_no"],
            "creditEnding": row["credit_ending"],
            "amount": round_money(row["amount"]),
            "hst": round_money(row["hst"]),
            "total": round_money(parse_amount(row["amount"]) + parse_amount(row["hst"])),
        }
        for row in lcbo_rows
    ]
    card_payments = [
        {
            "id": row["id"],
            "date": row["payment_date"],
            "purpose": row["purpose"],
            "amount": round_money(row["amount"]),
        }
        for row in payment_rows
    ]

    lcbo_amount = round_money(sum(parse_amount(row["amount"]) for row in lcbo_rows))
    lcbo_hst = round_money(sum(parse_amount(row["hst"]) for row in lcbo_rows))
    lcbo_total = round_money(lcbo_amount + lcbo_hst)
    paid_to_cards = round_money(sum(parse_amount(row["amount"]) for row in payment_rows))

    workflow = get_lcbo_workflow_row(db, store_id, year, month)
    timeline: list[dict[str, Any]] = []
    posted_payment: dict[str, Any] | None = None

    status = LCBO_STATUS_PENDING
    validated_amount = 0.0
    validated_at = None
    notes = ""
    posted_at = None
    posted_resource = None
    posted_record_id = None
    posted_payment_type = None
    workflow_id = None

    if workflow is not None:
        workflow_id = workflow["id"]
        status = workflow["status"]
        validated_amount = round_money(workflow["validated_amount"])
        validated_at = workflow["validated_at"]
        notes = workflow["notes"] or ""
        posted_at = workflow["posted_at"]
        posted_resource = workflow["posted_resource"]
        posted_record_id = workflow["posted_record_id"]
        posted_payment_type = workflow["posted_payment_type"]

        event_rows = db.execute(
            """
            SELECT id, event_type, from_status, to_status, event_amount, event_note, created_at
              FROM lcbo_monthly_workflow_events
             WHERE workflow_id = ?
             ORDER BY id DESC
            """,
            (workflow["id"],),
        ).fetchall()
        timeline = [
            {
                "id": event["id"],
                "eventType": event["event_type"],
                "fromStatus": event["from_status"],
                "toStatus": event["to_status"],
                "eventAmount": round_money(event["event_amount"]),
                "eventNote": event["event_note"],
                "createdAt": event["created_at"],
            }
            for event in event_rows
        ]

        if posted_resource and posted_record_id:
            if posted_resource == "cash-payments":
                payment_row = db.execute(
                    """
                    SELECT id, payment_date, vendor_name, amount, hst, source_type, source_ref_id, source_note
                      FROM cash_payments
                     WHERE id = ?
                    """,
                    (posted_record_id,),
                ).fetchone()
            elif posted_resource == "bank-payments":
                payment_row = db.execute(
                    """
                    SELECT id, payment_date, vendor_name, amount, hst, source_type, source_ref_id, source_note
                      FROM bank_payments
                     WHERE id = ?
                    """,
                    (posted_record_id,),
                ).fetchone()
            else:
                payment_row = None

            if payment_row is not None:
                posted_payment = {
                    "resource": posted_resource,
                    "recordId": payment_row["id"],
                    "date": payment_row["payment_date"],
                    "vendorName": payment_row["vendor_name"],
                    "amount": round_money(payment_row["amount"]),
                    "hst": round_money(payment_row["hst"]),
                    "total": round_money(parse_amount(payment_row["amount"]) + parse_amount(payment_row["hst"])),
                    "sourceType": payment_row["source_type"],
                    "sourceRefId": payment_row["source_ref_id"],
                    "sourceNote": payment_row["source_note"],
                }

    return {
        "store": store,
        "period": {
            "year": year,
            "month": month,
            "key": period_key,
            "label": month_label(year, month),
        },
        "lcboRecords": lcbo_records,
        "creditCardPayments": card_payments,
        "totals": {
            "lcboAmount": lcbo_amount,
            "lcboHst": lcbo_hst,
            "lcboTotal": lcbo_total,
            "paidToCreditCards": paid_to_cards,
            "difference": round_money(lcbo_total - paid_to_cards),
            "recommendedValidatedAmount": lcbo_total,
        },
        "workflow": {
            "id": workflow_id,
            "status": status,
            "statusLabel": lcbo_status_label(status),
            "validatedAmount": validated_amount,
            "validatedAt": validated_at,
            "notes": notes,
            "postedAt": posted_at,
            "postedResource": posted_resource,
            "postedRecordId": posted_record_id,
            "postedPaymentType": posted_payment_type,
            "postedPayment": posted_payment,
        },
        "timeline": timeline,
        "actions": {
            "canValidate": status != LCBO_STATUS_POSTED,
            "canPost": status == LCBO_STATUS_VALIDATED and validated_amount > 0,
            "canReverse": status == LCBO_STATUS_POSTED,
        },
    }


def fetch_rows(
    table: str,
    columns: list[str],
    *,
    where_clauses: list[str] | None = None,
    params: list[Any] | tuple[Any, ...] | None = None,
    order_by: str | None = None,
) -> list[dict[str, Any]]:
    query = f"SELECT {', '.join(columns)} FROM {table}"
    if where_clauses:
        query += f" WHERE {' AND '.join(where_clauses)}"
    if order_by:
        query += f" ORDER BY {order_by}"
    rows = get_db().execute(query, params or []).fetchall()
    return [dict(row) for row in rows]


def fetch_store_rows(
    table: str,
    date_col: str,
    columns: list[str],
    store_id: int,
    *,
    year: int | None = None,
    month: int | None = None,
    quarter: int | None = None,
) -> list[dict[str, Any]]:
    clauses = ["store_id = ?"]
    params: list[Any] = [store_id]
    period_clauses, period_params = build_period_filters(date_col, year, month, quarter)
    clauses.extend(period_clauses)
    params.extend(period_params)
    return fetch_rows(
        table,
        columns,
        where_clauses=clauses,
        params=params,
        order_by=f"{date_col} ASC, id ASC",
    )


def extract_month(date_value: str) -> int:
    return int(date_value[5:7])


def filter_rows_by_month(rows: list[dict[str, Any]], date_key: str, month: int) -> list[dict[str, Any]]:
    return [row for row in rows if row.get(date_key) and extract_month(str(row[date_key])) == month]


def sum_values(rows: list[dict[str, Any]], field: str) -> float:
    return sum(parse_amount(row.get(field)) for row in rows)


def build_store_dataset(store_id: int, year: int, month: int | None = None) -> dict[str, Any]:
    store = dict(get_store_row(store_id))
    return {
        "store": store,
        "year": year,
        "month": month,
        "sales_entries": fetch_store_rows(
            "daily_sales",
            "entry_date",
            ["entry_date", "total", "sales", "hst", "online", "instant", "cc", "gc", "non_add", "mc", "visa", "debit", "cash"],
            store_id,
            year=year,
            month=month,
        ),
        "lottery_records": fetch_store_rows(
            "lottery_records",
            "lottery_date",
            ["lottery_date", "lottery_payment", "lottery_income"],
            store_id,
            year=year,
            month=month,
        ),
        "cash_payments": fetch_store_rows(
            "cash_payments",
            "payment_date",
            ["payment_date", "vendor_name", "amount", "hst"],
            store_id,
            year=year,
            month=month,
        ),
        "bank_payments": fetch_store_rows(
            "bank_payments",
            "payment_date",
            ["payment_date", "vendor_name", "amount", "hst", "chq"],
            store_id,
            year=year,
            month=month,
        ),
        "expenses": fetch_store_rows(
            "expenses",
            "expense_date",
            ["expense_date", "vendor_name", "amount", "hst"],
            store_id,
            year=year,
            month=month,
        ),
        "salaries": fetch_store_rows(
            "salaries",
            "salary_date",
            ["salary_date", "employee", "amount"],
            store_id,
            year=year,
            month=month,
        ),
        "other_income": fetch_store_rows(
            "other_income",
            "income_date",
            ["income_date", "vendor_name", "amount"],
            store_id,
            year=year,
            month=month,
        ),
        "lcbo_entries": fetch_store_rows(
            "lcbo_entries",
            "entry_date",
            ["entry_date", "vendor_name", "invoice_no", "credit_ending", "amount", "hst"],
            store_id,
            year=year,
            month=month,
        ),
        "credit_card_payments": fetch_store_rows(
            "credit_card_payments",
            "payment_date",
            ["payment_date", "purpose", "amount"],
            store_id,
            year=year,
            month=month,
        ),
    }


def filter_dataset_by_month(dataset: dict[str, Any], month: int) -> dict[str, Any]:
    filtered = {"store": dataset["store"], "year": dataset["year"], "month": month}
    for key, date_key in DATASET_DATE_FIELDS.items():
        filtered[key] = filter_rows_by_month(dataset[key], date_key, month)
    return filtered


def build_period_summary(dataset: dict[str, Any]) -> dict[str, Any]:
    sales_entries = dataset["sales_entries"]
    lottery_records = dataset["lottery_records"]
    cash_payments = dataset["cash_payments"]
    bank_payments = dataset["bank_payments"]
    expenses = dataset["expenses"]
    salaries = dataset["salaries"]
    other_income = dataset["other_income"]
    lcbo_entries = dataset["lcbo_entries"]
    credit_card_payments = dataset["credit_card_payments"]

    pos = sum_values(sales_entries, "total")
    sales = sum_values(sales_entries, "sales")
    hst = sum_values(sales_entries, "hst")
    online = sum_values(sales_entries, "online")
    instant = sum_values(sales_entries, "instant")
    cc = sum_values(sales_entries, "cc")
    gc = sum_values(sales_entries, "gc")
    non_add = sum_values(sales_entries, "non_add")
    mc = sum_values(sales_entries, "mc")
    visa = sum_values(sales_entries, "visa")
    debit = sum_values(sales_entries, "debit")
    cash_deposits = sum_values(sales_entries, "cash")

    interac = mc + visa + debit
    lottery_payment = sum_values(lottery_records, "lottery_payment")
    lottery_income = sum_values(lottery_records, "lottery_income")
    other_income_total = sum_values(other_income, "amount")

    cash_payment_total = sum_values(cash_payments, "amount")
    bank_payment_total = sum_values(bank_payments, "amount")
    expense_total = sum_values(expenses, "amount")
    salary_total = sum_values(salaries, "amount")
    total_payouts = cash_payment_total + bank_payment_total + expense_total + salary_total

    lcbo_amount = sum_values(lcbo_entries, "amount")
    lcbo_hst = sum_values(lcbo_entries, "hst")
    lcbo_total = lcbo_amount + lcbo_hst
    credit_card_payment_total = sum_values(credit_card_payments, "amount")

    sales_total = pos + hst + online + instant + cc + gc
    income_total = interac + cash_deposits
    income_total_extended = income_total + lottery_income + other_income_total
    operating_difference = sales_total - total_payouts - income_total_extended
    balance_net = sales_total - (cash_payment_total + salary_total)
    balance_total_amount = interac + cash_deposits
    balance_difference = balance_total_amount - sales_total
    payout_pct = (total_payouts / pos) * 100 if abs(pos) > 0.009 else 0.0

    return {
        "pos": pos,
        "sales": sales,
        "hst": hst,
        "online": online,
        "instant": instant,
        "cc": cc,
        "gc": gc,
        "nonAdd": non_add,
        "mc": mc,
        "visa": visa,
        "debit": debit,
        "cashDeposits": cash_deposits,
        "interac": interac,
        "lotteryPayment": lottery_payment,
        "lotteryIncome": lottery_income,
        "otherIncome": other_income_total,
        "salesTotal": sales_total,
        "incomeTotal": income_total,
        "incomeTotalExtended": income_total_extended,
        "cashPayments": cash_payment_total,
        "bankPayments": bank_payment_total,
        "expenses": expense_total,
        "salaries": salary_total,
        "totalPayouts": total_payouts,
        "operatingDifference": operating_difference,
        "balanceNet": balance_net,
        "balanceTotalAmount": balance_total_amount,
        "balanceDifference": balance_difference,
        "payoutPct": payout_pct,
        "lcboAmount": lcbo_amount,
        "lcboHst": lcbo_hst,
        "lcboTotal": lcbo_total,
        "creditCardPayments": credit_card_payment_total,
    }


def build_annual_summary(monthly_summaries: list[dict[str, Any]]) -> dict[str, Any]:
    numeric_fields = [
        "pos",
        "sales",
        "hst",
        "online",
        "instant",
        "cc",
        "gc",
        "nonAdd",
        "mc",
        "visa",
        "debit",
        "cashDeposits",
        "interac",
        "lotteryPayment",
        "lotteryIncome",
        "otherIncome",
        "salesTotal",
        "incomeTotal",
        "incomeTotalExtended",
        "cashPayments",
        "bankPayments",
        "expenses",
        "salaries",
        "totalPayouts",
        "operatingDifference",
        "balanceNet",
        "balanceTotalAmount",
        "balanceDifference",
        "lcboAmount",
        "lcboHst",
        "lcboTotal",
        "creditCardPayments",
    ]
    totals = {"monthLabel": "Annual Total"}
    for field in numeric_fields:
        totals[field] = sum(parse_amount(item.get(field)) for item in monthly_summaries)
    totals["payoutPct"] = (totals["totalPayouts"] / totals["pos"]) * 100 if abs(totals["pos"]) > 0.009 else 0.0
    return totals


def sanitize_filename(value: str) -> str:
    safe = "".join(char.lower() if char.isalnum() else "-" for char in value).strip("-")
    while "--" in safe:
        safe = safe.replace("--", "-")
    return safe or "store"


def write_sheet_title(ws, row_index: int, title: str, width: int) -> int:
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=max(width, 1))
    cell = ws.cell(row=row_index, column=1, value=title)
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="left")
    return row_index + 1


def write_table(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    numeric_columns: set[int] | None = None,
    highlight_last_row: bool = False,
) -> int:
    numeric_columns = numeric_columns or set()
    row_index = write_sheet_title(ws, start_row, title, len(headers))

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    total_fill = PatternFill("solid", fgColor="E0F2FE")
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_index, column=col_index, value=header)
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row_index += 1

    if not rows:
        ws.cell(row=row_index, column=1, value="No records")
        return row_index + 2

    for item_index, values in enumerate(rows):
        is_total_row = highlight_last_row and item_index == len(rows) - 1
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if col_index in numeric_columns:
                cell.number_format = EXCEL_NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
            if is_total_row:
                cell.font = Font(bold=True)
                cell.fill = total_fill
        row_index += 1
    return row_index + 1


def autosize_columns(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 12), 24)


def export_month_sheet_name(month_number: int) -> str:
    return EXPORT_MONTH_CODES[month_number - 1]


def set_export_column_widths(ws, widths: dict[str, float]) -> None:
    for column_name, width in widths.items():
        ws.column_dimensions[column_name].width = width


def style_export_cell(
    cell,
    *,
    fill: PatternFill | None = None,
    bold: bool = False,
    color: str = "0F172A",
    align: str = "center",
    date_value: bool = False,
    numeric_value: bool = False,
    percent_value: bool = False,
) -> None:
    cell.border = EXPORT_BORDER
    cell.font = Font(bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fill is not None:
        cell.fill = fill
    if date_value and isinstance(cell.value, (date, datetime)):
        cell.number_format = EXCEL_DATE_FORMAT
    formula_number = isinstance(cell.value, str) and cell.value.startswith("=") and align == "right"
    if (numeric_value or formula_number) and cell.value not in (None, ""):
        cell.number_format = EXCEL_NUMBER_FORMAT
    if percent_value and cell.value not in (None, ""):
        cell.number_format = EXCEL_PERCENT_FORMAT


def fill_export_range(
    ws,
    cell_range: str,
    *,
    fill: PatternFill | None = None,
    bold: bool = False,
    color: str = "0F172A",
    align: str = "center",
) -> None:
    for row in ws[cell_range]:
        for cell in row:
            style_export_cell(cell, fill=fill, bold=bold, color=color, align=align)


def merge_export_title(ws, cell_range: str, value: str, *, fill: PatternFill | None = None, color: str = "FFFFFF") -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    style_export_cell(cell, fill=fill or EXPORT_TITLE_FILL, bold=True, color=color, align="center")


def amount_total(rows: list[dict[str, Any]]) -> float:
    return sum(parse_amount(row.get("amount")) for row in rows)


def hst_total(rows: list[dict[str, Any]]) -> float:
    return sum(parse_amount(row.get("hst")) for row in rows)


def amount_with_hst_total(rows: list[dict[str, Any]]) -> float:
    return sum(parse_amount(row.get("amount")) + parse_amount(row.get("hst")) for row in rows)


def excel_date(value: str | None) -> datetime | None:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d")


def display_amount(value: Any) -> float | None:
    amount = parse_amount(value)
    return None if abs(amount) < 0.000001 else amount


def summarize_vendor_payment_rows(rows: list[list[Any]]) -> list[list[Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row in rows:
        vendor_name = normalize_text(row[1]) or "Unknown"
        key = vendor_name.casefold()
        if key not in grouped:
            grouped[key] = {"vendor_name": vendor_name, "amount": 0.0, "hst": 0.0}
            order.append(key)
        grouped[key]["amount"] += parse_amount(row[2])
        grouped[key]["hst"] += parse_amount(row[3])
    return [
        [
            None,
            grouped[key]["vendor_name"],
            display_amount(round_money(grouped[key]["amount"])),
            display_amount(round_money(grouped[key]["hst"])),
        ]
        for key in order
    ]


def build_export_period_summary(dataset: dict[str, Any]) -> dict[str, Any]:
    sales_entries = dataset["sales_entries"]
    lottery_records = dataset["lottery_records"]
    cash_payments = dataset["cash_payments"]
    bank_payments = dataset["bank_payments"]
    expenses = dataset["expenses"]
    salaries = dataset["salaries"]
    other_income = dataset["other_income"]

    sales_total = sum_values(sales_entries, "total")
    sales = sum_values(sales_entries, "sales")
    hst = sum_values(sales_entries, "hst")
    online = sum_values(sales_entries, "online")
    instant = sum_values(sales_entries, "instant")
    cc = sum_values(sales_entries, "cc")
    gc = sum_values(sales_entries, "gc")
    non_add = sum_values(sales_entries, "non_add")
    mc = sum_values(sales_entries, "mc")
    visa = sum_values(sales_entries, "visa")
    debit = sum_values(sales_entries, "debit")
    cash = sum_values(sales_entries, "cash")

    interac = mc + visa + debit
    lottery_payment = sum_values(lottery_records, "lottery_payment")
    lottery_income = sum_values(lottery_records, "lottery_income")
    other_income_total = amount_total(other_income)

    cash_amount = amount_total(cash_payments)
    cash_hst = hst_total(cash_payments)
    cash_total = amount_with_hst_total(cash_payments)
    bank_amount = amount_total(bank_payments)
    bank_hst = hst_total(bank_payments)
    bank_total = amount_with_hst_total(bank_payments)
    expense_amount = amount_total(expenses)
    expense_hst = hst_total(expenses)
    expense_total = amount_with_hst_total(expenses)
    salary_total = amount_total(salaries)

    sales_rollup_total = sales_total + online + instant + cc + gc
    income_total = interac + cash
    difference = sales_rollup_total - cash_total - income_total
    balance_net = sales_rollup_total - cash_total - salary_total
    total_amount = interac + cash
    balance_difference = total_amount - balance_net
    total_payouts = cash_total + bank_total
    payout_pct = (total_payouts / sales_total) * 100 if abs(sales_total) > 0.009 else 0.0

    return {
        "salesTotal": sales_total,
        "sales": sales,
        "hst": hst,
        "online": online,
        "instant": instant,
        "cc": cc,
        "gc": gc,
        "ccPlusGc": cc + gc,
        "nonAdd": non_add,
        "mc": mc,
        "visa": visa,
        "debit": debit,
        "cash": cash,
        "interac": interac,
        "lotteryPayment": lottery_payment,
        "lotteryIncome": lottery_income,
        "otherIncome": other_income_total,
        "cashPaymentsAmount": cash_amount,
        "cashPaymentsHst": cash_hst,
        "cashPaymentsTotal": cash_total,
        "bankPaymentsAmount": bank_amount,
        "bankPaymentsHst": bank_hst,
        "bankPaymentsTotal": bank_total,
        "expensesAmount": expense_amount,
        "expensesHst": expense_hst,
        "expensesTotal": expense_total,
        "salaries": salary_total,
        "salesRollupTotal": sales_rollup_total,
        "incomeTotal": income_total,
        "difference": difference,
        "balanceNet": balance_net,
        "totalAmount": total_amount,
        "balanceDifference": balance_difference,
        "totalPayouts": total_payouts,
        "payoutPct": payout_pct,
    }


def write_monthly_side_table(
    ws,
    start_row: int,
    start_col: int,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    date_columns: set[int] | None = None,
    numeric_columns: set[int] | None = None,
) -> None:
    end_col = start_col + len(headers) - 1
    merge_export_title(
        ws,
        f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}",
        title,
        fill=EXPORT_GROUP_FILL,
        color="0F172A",
    )
    header_row = start_row + 1
    for index, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=header_row, column=index, value=header)
        style_export_cell(cell, fill=EXPORT_HEADER_FILL, bold=True)

    date_columns = date_columns or set()
    numeric_columns = numeric_columns or set()
    row_index = header_row + 1
    for values in rows:
        for offset, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=start_col + offset - 1, value=value)
            style_export_cell(
                cell,
                align="left" if offset not in numeric_columns else "right",
                date_value=offset in date_columns,
                numeric_value=offset in numeric_columns,
            )
        row_index += 1


def build_summary_sheet(workbook: Workbook, store_name: str, year: int, monthly_summaries: list[dict[str, Any]]) -> None:
    ws = workbook.active
    ws.title = "Summary"
    set_export_column_widths(
        ws,
        {"A": 18, "B": 14, "C": 14, "D": 14, "E": 14, "F": 15, "G": 16, "H": 14, "I": 15, "J": 14, "K": 14, "L": 14},
    )

    merge_export_title(ws, "A1:K1", "Balance Sheet")
    ws["A2"] = "Month"
    ws["A3"] = ""
    style_export_cell(ws["A2"], fill=EXPORT_GROUP_FILL, bold=True)
    style_export_cell(ws["A3"], fill=EXPORT_HEADER_FILL, bold=True)
    merge_export_title(ws, "B2:F2", "Sales", fill=EXPORT_GROUP_FILL, color="0F172A")
    ws["G2"] = "Purchases"
    style_export_cell(ws["G2"], fill=EXPORT_GROUP_FILL, bold=True)
    merge_export_title(ws, "H2:J2", "Income", fill=EXPORT_GROUP_FILL, color="0F172A")
    merge_export_title(ws, "K2:K3", "Difference", fill=EXPORT_GROUP_FILL, color="0F172A")

    top_headers = ["Total (POS+HST)", "Online", "Instant", "CC+GC", "Total Sales", "Cash Payments", "Interac", "Cash Deposits", "Total Income"]
    for offset, header in enumerate(top_headers, start=2):
        cell = ws.cell(row=3, column=offset, value=header)
        style_export_cell(cell, fill=EXPORT_HEADER_FILL, bold=True)

    for month_index, _summary in enumerate(monthly_summaries, start=4):
        month_number = month_index - 3
        month_name = MONTH_NAMES[month_index - 4]
        month_sheet = f"'{export_month_sheet_name(month_number)}'"
        ws.cell(row=month_index, column=1, value=month_name)
        style_export_cell(ws.cell(row=month_index, column=1), align="left")
        formulas = [
            f"={month_sheet}!B35",
            f"={month_sheet}!E35",
            f"={month_sheet}!F35",
            f"={month_sheet}!G35+{month_sheet}!H35",
            f"=SUM(B{month_index}:E{month_index})",
            f"=SUM({month_sheet}!U4:U999)+SUM({month_sheet}!V4:V999)",
            f"={month_sheet}!J35+{month_sheet}!K35+{month_sheet}!L35",
            f"={month_sheet}!M35",
            f"=H{month_index}+I{month_index}",
            f"=F{month_index}-G{month_index}-J{month_index}",
        ]
        for column_index, formula in enumerate(formulas, start=2):
            style_export_cell(
                ws.cell(row=month_index, column=column_index, value=formula),
                align="right",
            )

    total_row = 16
    ws.cell(row=total_row, column=1, value="Total")
    style_export_cell(ws.cell(row=total_row, column=1), fill=EXPORT_TOTAL_FILL, bold=True, align="left")
    for column_index in range(2, 12):
        formula = f"=SUM({get_column_letter(column_index)}4:{get_column_letter(column_index)}15)"
        style_export_cell(
            ws.cell(row=total_row, column=column_index, value=formula),
            fill=EXPORT_TOTAL_FILL,
            bold=True,
            numeric_value=False,
            align="right",
        )

    ws["A18"] = "Month"
    style_export_cell(ws["A18"], fill=EXPORT_GROUP_FILL, bold=True)
    merge_export_title(ws, "B18:C18", "Bank Payments", fill=EXPORT_GROUP_FILL, color="0F172A")
    style_export_cell(ws["D18"], fill=EXPORT_GROUP_FILL, bold=True)
    ws["D18"] = "Cash Salaries"
    merge_export_title(ws, "E18:F18", "Lottery", fill=EXPORT_GROUP_FILL, color="0F172A")
    ws["G18"] = "Other Income"
    style_export_cell(ws["G18"], fill=EXPORT_GROUP_FILL, bold=True)

    lower_headers = ["Bank Payments", "Utilities", "Cash Salaries", "Lotto Payments", "Lotto Income", "Other Income"]
    for offset, header in enumerate(lower_headers, start=2):
        cell = ws.cell(row=19, column=offset, value=header)
        style_export_cell(cell, fill=EXPORT_HEADER_FILL, bold=True)

    for row_index, _summary in enumerate(monthly_summaries, start=20):
        month_number = row_index - 19
        month_name = MONTH_NAMES[row_index - 20]
        month_sheet = f"'{export_month_sheet_name(month_number)}'"
        ws.cell(row=row_index, column=1, value=month_name)
        style_export_cell(ws.cell(row=row_index, column=1), align="left")
        formulas = [
            f"=SUM({month_sheet}!Y4:Y999)+SUM({month_sheet}!Z4:Z999)",
            f"=SUM({month_sheet}!L40:L999)+SUM({month_sheet}!M40:M999)",
            f"=SUM({month_sheet}!P40:P999)",
            f"={month_sheet}!N35",
            f"={month_sheet}!O35",
            f"=SUM({month_sheet}!S40:S999)",
        ]
        for column_index, formula in enumerate(formulas, start=2):
            style_export_cell(
                ws.cell(row=row_index, column=column_index, value=formula),
                align="right",
            )

    total_row = 32
    ws.cell(row=total_row, column=1, value="Total")
    style_export_cell(ws.cell(row=total_row, column=1), fill=EXPORT_TOTAL_FILL, bold=True, align="left")
    for column_index in range(2, 8):
        formula = f"=SUM({get_column_letter(column_index)}20:{get_column_letter(column_index)}31)"
        style_export_cell(
            ws.cell(row=total_row, column=column_index, value=formula),
            fill=EXPORT_TOTAL_FILL,
            bold=True,
            align="right",
        )

    ws["A35"] = "Calculation 1"
    style_export_cell(ws["A35"], bold=True, align="left")
    merge_export_title(ws, "B35:L35", "Quarterly Report", fill=EXPORT_TITLE_FILL)
    merge_export_title(ws, "B36:F36", "SALE", fill=EXPORT_GROUP_FILL, color="0F172A")
    merge_export_title(ws, "G36:K36", "Purchases", fill=EXPORT_GROUP_FILL, color="0F172A")
    ws["L36"] = "Difference"
    style_export_cell(ws["L36"], fill=EXPORT_GROUP_FILL, bold=True)

    quarterly_headers = [
        "Sales",
        "CC+GC",
        "Lottery Income",
        "Other Income",
        "Total Sales",
        "Cash Payments",
        "Bank Payments",
        "Utilities",
        "Salaries",
        "Total Payments",
    ]
    for column_index, header in enumerate(quarterly_headers, start=2):
        style_export_cell(ws.cell(row=37, column=column_index, value=header), fill=EXPORT_HEADER_FILL, bold=True)

    quarter_labels = ["Q1 (JAN - MAR)", "Q2 (APR - JUN)", "Q3 (JUL - SEP)", "Q4 (OCT - DEC)"]
    quarter_ranges = [(4, 6), (7, 9), (10, 12), (13, 15)]
    for row_index, (label, (start_row, end_row)) in enumerate(zip(quarter_labels, quarter_ranges), start=38):
        ws.cell(row=row_index, column=1, value=label)
        style_export_cell(ws.cell(row=row_index, column=1), align="left")
        formulas = [
            f"=SUM(B{start_row}:B{end_row})",
            f"=SUM(E{start_row}:E{end_row})",
            f"=SUM(F{start_row + 16}:F{end_row + 16})",
            f"=SUM(G{start_row + 16}:G{end_row + 16})",
            f"=SUM(B{row_index}:E{row_index})",
            f"=SUM(G{start_row}:G{end_row})",
            f"=SUM(B{start_row + 16}:B{end_row + 16})",
            f"=SUM(C{start_row + 16}:C{end_row + 16})",
            f"=SUM(D{start_row + 16}:D{end_row + 16})",
            f"=SUM(G{row_index}:J{row_index})",
            f"=F{row_index}-K{row_index}",
        ]
        for column_index, formula in enumerate(formulas, start=2):
            style_export_cell(ws.cell(row=row_index, column=column_index, value=formula), align="right")

    total_row = 42
    ws.cell(row=total_row, column=1, value="Total")
    style_export_cell(ws.cell(row=total_row, column=1), fill=EXPORT_TOTAL_FILL, bold=True, align="left")
    for column_index in range(2, 13):
        formula = f"=SUM({get_column_letter(column_index)}38:{get_column_letter(column_index)}41)"
        style_export_cell(
            ws.cell(row=total_row, column=column_index, value=formula),
            fill=EXPORT_TOTAL_FILL,
            bold=True,
            align="right",
        )


def build_hst_sheet(workbook: Workbook, monthly_summaries: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet(title="HST")
    set_export_column_widths(
        ws,
        {"A": 18, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 15, "H": 15, "I": 15, "J": 14, "K": 14},
    )

    merge_export_title(ws, "A1:K1", "Balance Sheet")
    ws["A2"] = "Month"
    style_export_cell(ws["A2"], fill=EXPORT_GROUP_FILL, bold=True)
    merge_export_title(ws, "B2:F2", "Sales", fill=EXPORT_GROUP_FILL, color="0F172A")
    merge_export_title(ws, "G2:J2", "Purchase", fill=EXPORT_GROUP_FILL, color="0F172A")
    merge_export_title(ws, "K2:K3", "Expenses", fill=EXPORT_GROUP_FILL, color="0F172A")

    top_headers = ["Sales", "Online", "Instant", "CC+GC", "Total", "Cash Payments", "Bank Payments", "Lotto Payments", "Total"]
    for offset, header in enumerate(top_headers, start=2):
        style_export_cell(ws.cell(row=3, column=offset, value=header), fill=EXPORT_HEADER_FILL, bold=True)

    for row_index, _summary in enumerate(monthly_summaries, start=4):
        month_number = row_index - 3
        month_name = MONTH_NAMES[row_index - 4]
        month_sheet = f"'{export_month_sheet_name(month_number)}'"
        ws.cell(row=row_index, column=1, value=month_name)
        style_export_cell(ws.cell(row=row_index, column=1), align="left")
        formulas = [
            f"={month_sheet}!C35",
            f"={month_sheet}!E35",
            f"={month_sheet}!F35",
            f"={month_sheet}!G35+{month_sheet}!H35",
            f"=SUM(B{row_index}:E{row_index})",
            f"=SUM({month_sheet}!U4:U999)",
            f"=SUM({month_sheet}!Y4:Y999)",
            f"={month_sheet}!N35",
            f"=SUM(G{row_index}:I{row_index})",
            f"=SUM({month_sheet}!L40:L999)",
        ]
        for column_index, formula in enumerate(formulas, start=2):
            style_export_cell(ws.cell(row=row_index, column=column_index, value=formula), align="right")

    total_row = 16
    ws.cell(row=total_row, column=1, value="Total")
    style_export_cell(ws.cell(row=total_row, column=1), fill=EXPORT_TOTAL_FILL, bold=True, align="left")
    for column_index in range(2, 12):
        formula = f"=SUM({get_column_letter(column_index)}4:{get_column_letter(column_index)}15)"
        style_export_cell(ws.cell(row=total_row, column=column_index, value=formula), fill=EXPORT_TOTAL_FILL, bold=True, align="right")

    ws["A18"] = "Month"
    style_export_cell(ws["A18"], fill=EXPORT_GROUP_FILL, bold=True)
    merge_export_title(ws, "B18:F18", "Sales / Purchases", fill=EXPORT_GROUP_FILL, color="0F172A")
    ws["G18"] = "Expenses"
    style_export_cell(ws["G18"], fill=EXPORT_GROUP_FILL, bold=True)
    merge_export_title(ws, "H18:K18", "Income", fill=EXPORT_GROUP_FILL, color="0F172A")

    lower_headers = ["Sales - HST", "Cash Payments", "Bank Payments", "Other Payments", "Total", "Utilities-HST", "Interac", "Cash Deposits", "Lotto Income", "Other Income"]
    for offset, header in enumerate(lower_headers, start=2):
        style_export_cell(ws.cell(row=19, column=offset, value=header), fill=EXPORT_HEADER_FILL, bold=True)

    for row_index, _summary in enumerate(monthly_summaries, start=20):
        month_number = row_index - 19
        month_name = MONTH_NAMES[row_index - 20]
        month_sheet = f"'{export_month_sheet_name(month_number)}'"
        ws.cell(row=row_index, column=1, value=month_name)
        style_export_cell(ws.cell(row=row_index, column=1), align="left")
        formulas = [
            f"={month_sheet}!D35",
            f"=SUM({month_sheet}!V4:V999)",
            f"=SUM({month_sheet}!Z4:Z999)",
            f"=SUM({month_sheet}!P40:P999)",
            f"=C{row_index}+D{row_index}",
            f"=SUM({month_sheet}!M40:M999)",
            f"={month_sheet}!J35+{month_sheet}!K35+{month_sheet}!L35",
            f"={month_sheet}!M35",
            f"={month_sheet}!O35",
            f"=SUM({month_sheet}!S40:S999)",
        ]
        for column_index, formula in enumerate(formulas, start=2):
            style_export_cell(ws.cell(row=row_index, column=column_index, value=formula), align="right")

    total_row = 32
    ws.cell(row=total_row, column=1, value="Total")
    style_export_cell(ws.cell(row=total_row, column=1), fill=EXPORT_TOTAL_FILL, bold=True, align="left")
    for column_index in range(2, 12):
        formula = f"=SUM({get_column_letter(column_index)}20:{get_column_letter(column_index)}31)"
        style_export_cell(ws.cell(row=total_row, column=column_index, value=formula), fill=EXPORT_TOTAL_FILL, bold=True, align="right")


def build_month_sheet(workbook: Workbook, month_name: str, month_data: dict[str, Any], summary: dict[str, Any]) -> None:
    month_number = month_data["month"]
    ws = workbook.create_sheet(title=export_month_sheet_name(month_number))
    ws.freeze_panes = "A4"
    set_export_column_widths(
        ws,
        {
            "A": 14,
            "B": 12,
            "C": 12,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 12,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 12,
            "N": 12,
            "O": 12,
            "Q": 24,
            "R": 16,
            "S": 16,
            "T": 24,
            "U": 16,
            "V": 16,
            "W": 16,
            "X": 24,
            "Y": 16,
            "Z": 16,
        },
    )

    ws["A1"] = month_name
    style_export_cell(ws["A1"], fill=EXPORT_TITLE_FILL, bold=True, color="FFFFFF", align="left")
    merge_export_title(ws, "Q1:R1", "Monthly Balance Sheet", fill=EXPORT_TITLE_FILL)
    merge_export_title(ws, "S1:V1", "Cash Payments", fill=EXPORT_TITLE_FILL)
    merge_export_title(ws, "W1:Z1", "Bank Payments", fill=EXPORT_TITLE_FILL)
    merge_export_title(ws, "A2:I2", "Sales", fill=EXPORT_GROUP_FILL, color="0F172A")
    merge_export_title(ws, "J2:M2", "Income", fill=EXPORT_GROUP_FILL, color="0F172A")
    merge_export_title(ws, "N2:O2", "Lottery Info", fill=EXPORT_GROUP_FILL, color="0F172A")

    daily_headers = [
        "Date",
        "Total",
        "Sales",
        "HST",
        "Online",
        "Instant",
        "CC",
        "GC",
        "Non Add",
        "Master Card",
        "Visa",
        "Debit",
        "Daily Cash",
        "Payment",
        "Income",
    ]
    for column_index, header in enumerate(daily_headers, start=1):
        style_export_cell(ws.cell(row=3, column=column_index, value=header), fill=EXPORT_HEADER_FILL, bold=True)

    sales_by_date = {row["entry_date"]: row for row in month_data["sales_entries"]}
    lottery_by_date = {row["lottery_date"]: row for row in month_data["lottery_records"]}
    days_in_month = calendar.monthrange(month_data["year"], month_number)[1]

    for day_index in range(1, 32):
        row_index = 3 + day_index
        if day_index <= days_in_month:
            current_date = date(month_data["year"], month_number, day_index)
            date_key = current_date.isoformat()
            ws.cell(row=row_index, column=1, value=current_date)
            style_export_cell(ws.cell(row=row_index, column=1), date_value=True, align="left")
        else:
            date_key = None
        entry = sales_by_date.get(date_key or "", {})
        lottery = lottery_by_date.get(date_key or "", {})

        data_values = {
            2: display_amount(entry.get("total")),
            3: display_amount(entry.get("sales")),
            5: display_amount(entry.get("online")),
            6: display_amount(entry.get("instant")),
            7: display_amount(entry.get("cc")),
            8: display_amount(entry.get("gc")),
            9: display_amount(entry.get("non_add")),
            10: display_amount(entry.get("mc")),
            11: display_amount(entry.get("visa")),
            12: display_amount(entry.get("debit")),
            13: display_amount(entry.get("cash")),
            14: display_amount(lottery.get("lottery_payment")),
            15: display_amount(lottery.get("lottery_income")),
        }
        for column_index, value in data_values.items():
            if value is not None:
                style_export_cell(
                    ws.cell(row=row_index, column=column_index, value=value),
                    numeric_value=True,
                    align="right",
                )
            else:
                style_export_cell(ws.cell(row=row_index, column=column_index, value=None), align="right")
        ws.cell(row=row_index, column=4, value=f'=IF(OR(B{row_index}<>"",C{row_index}<>""),B{row_index}-C{row_index},"")')
        style_export_cell(ws.cell(row=row_index, column=4), align="right")

    total_row = 35
    ws.cell(row=total_row, column=1, value="Total")
    style_export_cell(ws.cell(row=total_row, column=1), fill=EXPORT_TOTAL_FILL, bold=True, align="left")
    for column_index in range(2, 16):
        formula = f"=SUM({get_column_letter(column_index)}4:{get_column_letter(column_index)}34)"
        style_export_cell(ws.cell(row=total_row, column=column_index, value=formula), fill=EXPORT_TOTAL_FILL, bold=True, align="right")

    balance_rows = {
        3: ("POS", "=C35"),
        4: ("HST", "=D35"),
        5: ("Online", "=E35"),
        6: ("Instant", "=F35"),
        7: ("CC", "=G35"),
        8: ("GC", "=H35"),
        10: ("Total Sales (A)", "=SUM(R3:R8)"),
        11: ("Net (A-B)=(C)", summary["balanceNet"]),
        12: ("Master Card (D)", "=J35"),
        13: ("Visa (E)", "=K35"),
        14: ("Debit (F)", "=L35"),
        15: ("Cash (G)", "=M35"),
        16: ("Total Amount (D+E+F+G)", "=SUM(R12:R15)"),
        17: ("Difference", "=R16-R11"),
        20: ("Percentage", None),
        21: ("Total Sales", "=B35"),
        22: ("Cash Payouts", "=SUM(U4:U999)+SUM(V4:V999)"),
        23: ("Bank Payouts", "=SUM(Y4:Y999)+SUM(Z4:Z999)"),
        24: ("Total Payouts", "=R22+R23"),
        25: ("Payout %", '=IF(R21=0,0,(R24/R21)*100)'),
    }
    for row_index, (label, value) in balance_rows.items():
        style_export_cell(ws.cell(row=row_index, column=17, value=label), fill=EXPORT_GROUP_FILL if row_index not in {11, 16, 17} else EXPORT_HIGHLIGHT_FILL, bold=True, align="left")
        value_cell = ws.cell(row=row_index, column=18, value=value)
        style_export_cell(
            value_cell,
            fill=EXPORT_GROUP_FILL if row_index not in {11, 16, 17} else EXPORT_HIGHLIGHT_FILL,
            bold=row_index in {10, 11, 16, 17, 24, 25},
            align="right",
            numeric_value=isinstance(value, (int, float)),
            percent_value=row_index == 25 and isinstance(value, (int, float)),
        )

    cash_rows = [
        [excel_date(row["payment_date"]), row["vendor_name"], display_amount(row["amount"]), display_amount(row["hst"])]
        for row in month_data["cash_payments"]
    ]
    bank_rows = [
        [excel_date(row["payment_date"]), row["vendor_name"], display_amount(row["amount"]), display_amount(row["hst"]), row["chq"]]
        for row in month_data["bank_payments"]
    ]
    expense_rows = [
        [excel_date(row["expense_date"]), row["vendor_name"], display_amount(row["amount"]), display_amount(row["hst"])]
        for row in month_data["expenses"]
    ]
    salary_rows = [
        [excel_date(row["salary_date"]), row["employee"], display_amount(row["amount"])]
        for row in month_data["salaries"]
    ]
    other_income_rows = [
        [excel_date(row["income_date"]), row["vendor_name"], display_amount(row["amount"])]
        for row in month_data["other_income"]
    ]
    cash_vendor_rows = summarize_vendor_payment_rows(cash_rows)
    bank_vendor_rows = summarize_vendor_payment_rows([[row[0], row[1], row[2], row[3]] for row in bank_rows])

    write_monthly_side_table(
        ws,
        2,
        19,
        "Cash Payments by Vendor",
        ["Date", "Name", "Amount", "HST"],
        cash_vendor_rows,
        date_columns={1},
        numeric_columns={3, 4},
    )
    write_monthly_side_table(
        ws,
        2,
        23,
        "Bank Payments by Vendor",
        ["Date", "Name", "Amount", "HST"],
        bank_vendor_rows,
        date_columns={1},
        numeric_columns={3, 4},
    )

    merge_export_title(ws, "A37:S37", "Purchases", fill=EXPORT_TITLE_FILL)
    write_monthly_side_table(ws, 38, 1, "Cash Payments", ["Date", "Name", "Amount", "HST"], cash_rows, date_columns={1}, numeric_columns={3, 4})
    write_monthly_side_table(ws, 38, 5, "Bank Payments", ["Date", "Name", "Amount", "HST", "# CHQ"], bank_rows, date_columns={1}, numeric_columns={3, 4})
    write_monthly_side_table(ws, 38, 10, "Expenses", ["Date", "Name", "Amount", "HST"], expense_rows, date_columns={1}, numeric_columns={3, 4})
    write_monthly_side_table(ws, 38, 14, "Salaries", ["Date", "Employee", "Amount"], salary_rows, date_columns={1}, numeric_columns={3})
    write_monthly_side_table(ws, 38, 17, "Other Income", ["Date", "Name", "Amount"], other_income_rows, date_columns={1}, numeric_columns={3})


def build_vendor_expense_sheet(workbook: Workbook, title: str, rows: list[sqlite3.Row | dict[str, Any]]) -> None:
    ws = workbook.create_sheet(title=title[:31])
    ws["A1"] = "Date"
    ws["B1"] = "Vendor"
    ws["C1"] = "Amount"
    ws["D1"] = "HST"
    for column in "ABCD":
        style_export_cell(ws[f"{column}1"], fill=EXPORT_HEADER_FILL, bold=True)
    set_export_column_widths(ws, {"A": 14, "B": 22, "C": 12, "D": 12})
    row_index = 2
    for row in rows:
        values = [
            excel_date(row["expense_date"]),
            row["vendor_name"],
            display_amount(row["amount"]),
            display_amount(row.get("hst")),
        ]
        for column_index, value in enumerate(values, start=1):
            style_export_cell(
                ws.cell(row=row_index, column=column_index, value=value),
                date_value=column_index == 1,
                numeric_value=column_index in {3, 4} and isinstance(value, (int, float)),
                align="left" if column_index in {1, 2} else "right",
            )
        row_index += 1


def build_resource_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]], *, date_columns: set[int], numeric_columns: set[int]) -> None:
    ws = workbook.create_sheet(title=title[:31])
    set_export_column_widths(ws, {get_column_letter(index): 16 for index in range(1, len(headers) + 1)})
    for column_index, header in enumerate(headers, start=1):
        style_export_cell(ws.cell(row=1, column=column_index, value=header), fill=EXPORT_HEADER_FILL, bold=True)
    row_index = 2
    for values in rows:
        for column_index, value in enumerate(values, start=1):
            style_export_cell(
                ws.cell(row=row_index, column=column_index, value=value),
                date_value=column_index in date_columns,
                numeric_value=column_index in numeric_columns and isinstance(value, (int, float)),
                align="left" if column_index not in numeric_columns else "right",
            )
        row_index += 1


def parse_excel_number(value: Any, cell_ref: str) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return round_money(value)
    text = normalize_text(value)
    if text == "":
        return 0.0
    if text.startswith("="):
        return 0.0
    cleaned = text.replace(",", "").replace("$", "")
    try:
        amount = Decimal(cleaned).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return float(amount)
    except InvalidOperation:
        abort(400, description=f"Invalid number in workbook at {cell_ref}: {text}")


def parse_excel_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
        return "" if text.startswith("=") else text
    if isinstance(value, bool):
        return ""
    return str(value).strip()


def parse_excel_date(value: Any, cell_ref: str) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = parse_excel_text(value)
    if text == "":
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    abort(400, description=f"Invalid date in workbook at {cell_ref}: {text}")


def parse_export_month_sheet(ws, year: int, month: int) -> dict[str, list[dict[str, Any]]]:
    days_in_month = calendar.monthrange(year, month)[1]
    month_start_date = f"{year:04d}-{month:02d}-01"
    daily_sales_rows: list[dict[str, Any]] = []
    lottery_rows: list[dict[str, Any]] = []
    cash_rows: list[dict[str, Any]] = []
    bank_rows: list[dict[str, Any]] = []
    expense_rows: list[dict[str, Any]] = []
    salary_rows: list[dict[str, Any]] = []
    other_income_rows: list[dict[str, Any]] = []

    for day in range(1, days_in_month + 1):
        row_index = 3 + day
        date_value = f"{year:04d}-{month:02d}-{day:02d}"
        total = parse_excel_number(ws.cell(row=row_index, column=2).value, f"{ws.title}!B{row_index}")
        sales = parse_excel_number(ws.cell(row=row_index, column=3).value, f"{ws.title}!C{row_index}")
        raw_hst = ws.cell(row=row_index, column=4).value
        if isinstance(raw_hst, str) and normalize_text(raw_hst).startswith("="):
            hst = round_money(total - sales) if abs(total) > 0.000001 or abs(sales) > 0.000001 else 0.0
        else:
            hst = parse_excel_number(raw_hst, f"{ws.title}!D{row_index}")
            if abs((sales + hst) - total) > 0.02 and (abs(total) > 0.000001 or abs(sales) > 0.000001):
                hst = round_money(total - sales)
        online = parse_excel_number(ws.cell(row=row_index, column=5).value, f"{ws.title}!E{row_index}")
        instant = parse_excel_number(ws.cell(row=row_index, column=6).value, f"{ws.title}!F{row_index}")
        cc = parse_excel_number(ws.cell(row=row_index, column=7).value, f"{ws.title}!G{row_index}")
        gc = parse_excel_number(ws.cell(row=row_index, column=8).value, f"{ws.title}!H{row_index}")
        non_add = parse_excel_number(ws.cell(row=row_index, column=9).value, f"{ws.title}!I{row_index}")
        mc = parse_excel_number(ws.cell(row=row_index, column=10).value, f"{ws.title}!J{row_index}")
        visa = parse_excel_number(ws.cell(row=row_index, column=11).value, f"{ws.title}!K{row_index}")
        debit = parse_excel_number(ws.cell(row=row_index, column=12).value, f"{ws.title}!L{row_index}")
        cash = parse_excel_number(ws.cell(row=row_index, column=13).value, f"{ws.title}!M{row_index}")
        lottery_payment = parse_excel_number(ws.cell(row=row_index, column=14).value, f"{ws.title}!N{row_index}")
        lottery_income = parse_excel_number(ws.cell(row=row_index, column=15).value, f"{ws.title}!O{row_index}")

        has_daily = any(
            abs(value) > 0.000001
            for value in (total, sales, hst, online, instant, cc, gc, non_add, mc, visa, debit, cash)
        )
        if has_daily:
            daily_sales_rows.append(
                {
                    "entry_date": date_value,
                    "total": round_money(sales + hst),
                    "sales": round_money(sales),
                    "hst": round_money(hst),
                    "online": round_money(online),
                    "instant": round_money(instant),
                    "cc": round_money(cc),
                    "gc": round_money(gc),
                    "non_add": round_money(non_add),
                    "mc": round_money(mc),
                    "visa": round_money(visa),
                    "debit": round_money(debit),
                    "cash": round_money(cash),
                    "lottery_payment": round_money(lottery_payment),
                    "lottery_income": round_money(lottery_income),
                }
            )

        if abs(lottery_payment) > 0.000001 or abs(lottery_income) > 0.000001:
            lottery_rows.append(
                {
                    "lottery_date": date_value,
                    "lottery_payment": round_money(lottery_payment),
                    "lottery_income": round_money(lottery_income),
                }
            )

    top_cash_rows: list[dict[str, Any]] = []
    top_bank_rows: list[dict[str, Any]] = []
    for row_index in range(4, ws.max_row + 1):
        cash_vendor = parse_excel_text(ws.cell(row=row_index, column=20).value)
        cash_amount = parse_excel_number(ws.cell(row=row_index, column=21).value, f"{ws.title}!U{row_index}")
        cash_hst = parse_excel_number(ws.cell(row=row_index, column=22).value, f"{ws.title}!V{row_index}")
        cash_has_values = abs(cash_amount) > 0.000001 or abs(cash_hst) > 0.000001
        if cash_vendor == "" and cash_has_values:
            abort(400, description=f"Missing cash payment vendor at {ws.title}!T{row_index}")
        if cash_has_values:
            cash_date = parse_excel_date(ws.cell(row=row_index, column=19).value, f"{ws.title}!S{row_index}")
            top_cash_rows.append(
                {
                    "payment_date": cash_date or month_start_date,
                    "vendor_name": cash_vendor,
                    "amount": round_money(cash_amount),
                    "hst": round_money(cash_hst),
                }
            )

        bank_vendor = parse_excel_text(ws.cell(row=row_index, column=24).value)
        bank_amount = parse_excel_number(ws.cell(row=row_index, column=25).value, f"{ws.title}!Y{row_index}")
        bank_hst = parse_excel_number(ws.cell(row=row_index, column=26).value, f"{ws.title}!Z{row_index}")
        bank_has_values = abs(bank_amount) > 0.000001 or abs(bank_hst) > 0.000001
        if bank_vendor == "" and not bank_has_values:
            continue
        if bank_vendor == "" and bank_has_values:
            abort(400, description=f"Missing bank payment vendor at {ws.title}!X{row_index}")
        if not bank_has_values:
            continue
        bank_date = parse_excel_date(ws.cell(row=row_index, column=23).value, f"{ws.title}!W{row_index}")
        top_bank_rows.append(
            {
                "payment_date": bank_date or month_start_date,
                "vendor_name": bank_vendor,
                "amount": round_money(bank_amount),
                "hst": round_money(bank_hst),
                "chq": "",
            }
        )

    for row_index in range(40, ws.max_row + 1):
        cash_date = parse_excel_date(ws.cell(row=row_index, column=1).value, f"{ws.title}!A{row_index}")
        cash_vendor = parse_excel_text(ws.cell(row=row_index, column=2).value)
        cash_amount = parse_excel_number(ws.cell(row=row_index, column=3).value, f"{ws.title}!C{row_index}")
        cash_hst = parse_excel_number(ws.cell(row=row_index, column=4).value, f"{ws.title}!D{row_index}")
        if cash_vendor == "" and abs(cash_amount) > 0.000001:
            abort(400, description=f"Missing cash payment vendor at {ws.title}!B{row_index}")
        if cash_vendor and (abs(cash_amount) > 0.000001 or abs(cash_hst) > 0.000001):
            cash_rows.append(
                {
                    "payment_date": cash_date or month_start_date,
                    "vendor_name": cash_vendor,
                    "amount": round_money(cash_amount),
                    "hst": round_money(cash_hst),
                }
            )

        bank_date = parse_excel_date(ws.cell(row=row_index, column=5).value, f"{ws.title}!E{row_index}")
        bank_vendor = parse_excel_text(ws.cell(row=row_index, column=6).value)
        bank_amount = parse_excel_number(ws.cell(row=row_index, column=7).value, f"{ws.title}!G{row_index}")
        bank_hst = parse_excel_number(ws.cell(row=row_index, column=8).value, f"{ws.title}!H{row_index}")
        bank_chq = parse_excel_text(ws.cell(row=row_index, column=9).value)
        if bank_vendor == "" and abs(bank_amount) > 0.000001:
            abort(400, description=f"Missing bank payment vendor at {ws.title}!F{row_index}")
        if bank_vendor and (abs(bank_amount) > 0.000001 or abs(bank_hst) > 0.000001):
            bank_rows.append(
                {
                    "payment_date": bank_date or month_start_date,
                    "vendor_name": bank_vendor,
                    "amount": round_money(bank_amount),
                    "hst": round_money(bank_hst),
                    "chq": bank_chq,
                }
            )

    if not cash_rows and top_cash_rows:
        cash_rows.extend(top_cash_rows)
    if not bank_rows and top_bank_rows:
        bank_rows.extend(top_bank_rows)

        expense_date = parse_excel_date(ws.cell(row=row_index, column=10).value, f"{ws.title}!J{row_index}")
        expense_vendor = parse_excel_text(ws.cell(row=row_index, column=11).value)
        expense_amount = parse_excel_number(ws.cell(row=row_index, column=12).value, f"{ws.title}!L{row_index}")
        expense_hst = parse_excel_number(ws.cell(row=row_index, column=13).value, f"{ws.title}!M{row_index}")
        if expense_vendor == "" and abs(expense_amount) > 0.000001:
            abort(400, description=f"Missing expense vendor at {ws.title}!K{row_index}")
        if expense_vendor and (abs(expense_amount) > 0.000001 or abs(expense_hst) > 0.000001):
            expense_rows.append(
                {
                    "expense_date": expense_date or month_start_date,
                    "vendor_name": expense_vendor,
                    "amount": round_money(expense_amount),
                    "hst": round_money(expense_hst),
                }
            )

        salary_date = parse_excel_date(ws.cell(row=row_index, column=14).value, f"{ws.title}!N{row_index}")
        salary_employee = parse_excel_text(ws.cell(row=row_index, column=15).value)
        salary_amount = parse_excel_number(ws.cell(row=row_index, column=16).value, f"{ws.title}!P{row_index}")
        if salary_employee == "" and abs(salary_amount) > 0.000001:
            abort(400, description=f"Missing salary employee at {ws.title}!O{row_index}")
        if salary_employee and abs(salary_amount) > 0.000001:
            salary_rows.append(
                {
                    "salary_date": salary_date or month_start_date,
                    "employee": salary_employee,
                    "amount": round_money(salary_amount),
                }
            )

        income_date = parse_excel_date(ws.cell(row=row_index, column=17).value, f"{ws.title}!Q{row_index}")
        income_vendor = parse_excel_text(ws.cell(row=row_index, column=18).value)
        income_amount = parse_excel_number(ws.cell(row=row_index, column=19).value, f"{ws.title}!S{row_index}")
        if income_vendor == "" and abs(income_amount) > 0.000001:
            abort(400, description=f"Missing other income vendor at {ws.title}!R{row_index}")
        if income_vendor and abs(income_amount) > 0.000001:
            other_income_rows.append(
                {
                    "income_date": income_date or month_start_date,
                    "vendor_name": income_vendor,
                    "amount": round_money(income_amount),
                }
            )

    return {
        "daily_sales": daily_sales_rows,
        "lottery_records": lottery_rows,
        "cash_payments": cash_rows,
        "bank_payments": bank_rows,
        "expenses": expense_rows,
        "salaries": salary_rows,
        "other_income": other_income_rows,
    }


def parse_lcbo_import_sheet(ws, year: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        entry_date = parse_excel_date(ws.cell(row=row_index, column=1).value, f"{ws.title}!A{row_index}")
        vendor_name = parse_excel_text(ws.cell(row=row_index, column=2).value) or "LCBO"
        invoice_no = parse_excel_text(ws.cell(row=row_index, column=3).value)
        credit_ending = parse_excel_text(ws.cell(row=row_index, column=4).value)
        amount = parse_excel_number(ws.cell(row=row_index, column=5).value, f"{ws.title}!E{row_index}")
        hst = parse_excel_number(ws.cell(row=row_index, column=6).value, f"{ws.title}!F{row_index}")
        if not entry_date and abs(amount) <= 0.000001 and abs(hst) <= 0.000001 and invoice_no == "" and credit_ending == "":
            continue
        if not entry_date:
            abort(400, description=f"Missing LCBO date at {ws.title}!A{row_index}")
        if int(entry_date[:4]) != year:
            abort(400, description=f"LCBO date year must be {year}: {ws.title}!A{row_index}")
        rows.append(
            {
                "entry_date": entry_date,
                "vendor_name": vendor_name,
                "invoice_no": invoice_no,
                "credit_ending": credit_ending,
                "amount": round_money(amount),
                "hst": round_money(hst),
            }
        )
    return rows


def parse_credit_card_import_sheet(ws, year: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        payment_date = parse_excel_date(ws.cell(row=row_index, column=1).value, f"{ws.title}!A{row_index}")
        purpose = parse_excel_text(ws.cell(row=row_index, column=2).value)
        amount = parse_excel_number(ws.cell(row=row_index, column=3).value, f"{ws.title}!C{row_index}")
        if not payment_date and purpose == "" and abs(amount) <= 0.000001:
            continue
        if not payment_date:
            abort(400, description=f"Missing credit card payment date at {ws.title}!A{row_index}")
        if int(payment_date[:4]) != year:
            abort(400, description=f"Credit card payment year must be {year}: {ws.title}!A{row_index}")
        if purpose == "":
            abort(400, description=f"Missing credit card purpose at {ws.title}!B{row_index}")
        rows.append(
            {
                "payment_date": payment_date,
                "purpose": purpose,
                "amount": round_money(amount),
            }
        )
    return rows


def parse_annual_import_workbook(workbook, year: int) -> dict[str, Any]:
    monthly_data: dict[int, dict[str, list[dict[str, Any]]]] = {}
    for month_index in range(1, 13):
        sheet_name = export_month_sheet_name(month_index)
        if sheet_name in workbook.sheetnames:
            monthly_data[month_index] = parse_export_month_sheet(workbook[sheet_name], year, month_index)
    if not monthly_data:
        abort(400, description="Workbook does not contain JAN-DEC sheets in export format")

    lcbo_rows: list[dict[str, Any]] | None = None
    cc_rows: list[dict[str, Any]] | None = None
    if "LCBO" in workbook.sheetnames:
        lcbo_rows = parse_lcbo_import_sheet(workbook["LCBO"], year)
    if "Credit Cards" in workbook.sheetnames:
        cc_rows = parse_credit_card_import_sheet(workbook["Credit Cards"], year)
    return {"months": monthly_data, "lcbo_entries": lcbo_rows, "credit_card_payments": cc_rows}


def replace_month_rows(
    db: sqlite3.Connection,
    *,
    table: str,
    date_col: str,
    store_id: int,
    year: int,
    month: int,
    columns: list[str],
    rows: list[tuple[Any, ...]],
) -> None:
    period_key = month_period_key(year, month)
    db.execute(
        f"DELETE FROM {table} WHERE store_id = ? AND {month_filter_sql(date_col)} = ?",
        (store_id, period_key),
    )
    if not rows:
        return
    placeholders = ", ".join("?" for _ in columns)
    db.executemany(
        f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})",
        rows,
    )


def replace_year_rows(
    db: sqlite3.Connection,
    *,
    table: str,
    date_col: str,
    store_id: int,
    year: int,
    columns: list[str],
    rows: list[tuple[Any, ...]],
) -> None:
    db.execute(
        f"DELETE FROM {table} WHERE store_id = ? AND {year_filter_sql(date_col)} = ?",
        (store_id, f"{year:04d}"),
    )
    if not rows:
        return
    placeholders = ", ".join("?" for _ in columns)
    db.executemany(
        f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})",
        rows,
    )

@app.get("/")
def index():
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.get("/css/<path:filename>")
def serve_css(filename: str):
    return send_from_directory(FRONTEND_DIR / "css", filename)


@app.get("/js/<path:filename>")
def serve_js(filename: str):
    return send_from_directory(FRONTEND_DIR / "js", filename)


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/api/auth/status")
def auth_status():
    return jsonify(
        {
            "authenticated": is_authenticated(),
            "username": session.get("username") if is_authenticated() else None,
            "authKey": get_auth_key(),
        }
    )


@app.post("/api/auth/login")
def auth_login():
    payload = request.get_json(silent=True) or {}
    username = normalize_text(payload.get("username"))
    password = normalize_text(payload.get("password"))
    if username != MASTER_USERNAME or password != MASTER_PASSWORD:
        abort(401, description="Invalid username or password")

    session.permanent = True
    session["authenticated"] = True
    session["username"] = MASTER_USERNAME
    session["auth_key"] = secrets.token_urlsafe(24)
    return jsonify({"authenticated": True, "username": MASTER_USERNAME, "authKey": session["auth_key"]})


@app.post("/api/auth/logout")
def auth_logout():
    session.clear()
    session.modified = True
    return jsonify({"authenticated": False, "authKey": None})


@app.get("/api/stores")
def list_stores():
    rows = get_db().execute("SELECT id, name, created_at FROM stores ORDER BY lower(name), id").fetchall()
    return jsonify([dict(row) for row in rows])


@app.post("/api/stores")
def create_store():
    payload = request.get_json(silent=True) or {}
    name = normalize_text(payload.get("name"))
    if not name:
        abort(400, description="Store name is required")

    db = get_db()
    cursor = db.execute("INSERT INTO stores (name) VALUES (?)", (name,))
    db.commit()

    row = db.execute("SELECT id, name, created_at FROM stores WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.patch("/api/stores/<int:store_id>")
def update_store(store_id: int):
    payload = request.get_json(silent=True) or {}
    name = normalize_text(payload.get("name"))
    if not name:
        abort(400, description="Store name is required")

    db = get_db()
    cursor = db.execute("UPDATE stores SET name = ? WHERE id = ?", (name, store_id))
    if cursor.rowcount == 0:
        abort(404, description="Store not found")
    db.commit()

    row = db.execute("SELECT id, name, created_at FROM stores WHERE id = ?", (store_id,)).fetchone()
    return jsonify(dict(row))


@app.delete("/api/stores/<int:store_id>")
def delete_store(store_id: int):
    db = get_db()
    store_count = db.execute("SELECT COUNT(*) FROM stores").fetchone()[0]
    if store_count <= 1:
        abort(400, description="At least one store must remain")

    cursor = db.execute("DELETE FROM stores WHERE id = ?", (store_id,))
    if cursor.rowcount == 0:
        abort(404, description="Store not found")

    db.commit()
    return "", 204


@app.get("/api/stores/<int:store_id>/daily-sales/<entry_date>")
def get_daily_sales_entry(store_id: int, entry_date: str):
    ensure_store_exists(store_id)
    row = get_db().execute(
        """
        SELECT id, entry_date, total, sales, hst, online, instant, cc, gc, non_add,
               mc, visa, debit, cash, lottery_payment, lottery_income
          FROM daily_sales
         WHERE store_id = ? AND entry_date = ?
        """,
        (store_id, entry_date),
    ).fetchone()
    return jsonify(serialize_daily_sales(row) if row else None)


@app.put("/api/stores/<int:store_id>/daily-sales/<entry_date>")
def save_daily_sales_entry(store_id: int, entry_date: str):
    ensure_store_exists(store_id)
    payload = request.get_json(silent=True) or {}
    payload["date"] = entry_date
    require_fields(payload, ("date",))
    validate_sales_totals(payload)

    values = map_payload(payload, DAILY_SALES_MAPPING, {key for key in DAILY_SALES_MAPPING if key != "date"}, partial=False)
    columns = ["store_id", *values.keys()]
    placeholders = ", ".join("?" for _ in columns)
    update_columns = ", ".join(f"{column} = excluded.{column}" for column in values)

    db = get_db()
    db.execute(
        f"""
        INSERT INTO daily_sales ({", ".join(columns)})
        VALUES ({placeholders})
        ON CONFLICT(store_id, entry_date) DO UPDATE SET
            {update_columns},
            updated_at = CURRENT_TIMESTAMP
        """,
        (store_id, *values.values()),
    )
    db.commit()

    row = db.execute(
        """
        SELECT id, entry_date, total, sales, hst, online, instant, cc, gc, non_add,
               mc, visa, debit, cash, lottery_payment, lottery_income
          FROM daily_sales
         WHERE store_id = ? AND entry_date = ?
        """,
        (store_id, entry_date),
    ).fetchone()
    return jsonify(serialize_daily_sales(row))


@app.get("/api/stores/<int:store_id>/daily-sales")
def list_daily_sales(store_id: int):
    ensure_store_exists(store_id)
    where_sql, params = build_date_filters(store_id, "entry_date")
    rows = get_db().execute(
        f"""
        SELECT id, entry_date, total, sales, hst, online, instant, cc, gc, non_add,
               mc, visa, debit, cash, lottery_payment, lottery_income
          FROM daily_sales
         WHERE {where_sql}
         ORDER BY entry_date ASC, id ASC
        """,
        params,
    ).fetchall()
    return jsonify([serialize_daily_sales(row) for row in rows])


@app.get("/api/reports/store-comparison")
def store_comparison():
    today_value = date.today()
    year = request.args.get("year", type=int) or today_value.year

    stores = fetch_rows("stores", ["id", "name"], order_by="lower(name), id ASC")
    summaries: list[dict[str, Any]] = []
    for store in stores:
        dataset = build_store_dataset(store["id"], year)
        summary = build_period_summary(dataset)
        summary["storeId"] = store["id"]
        summary["storeName"] = store["name"]
        summaries.append(summary)

    numeric_fields = [
        "pos",
        "salesTotal",
        "interac",
        "cashDeposits",
        "lotteryIncome",
        "otherIncome",
        "incomeTotalExtended",
        "cashPayments",
        "bankPayments",
        "expenses",
        "salaries",
        "totalPayouts",
        "operatingDifference",
        "lcboTotal",
        "creditCardPayments",
    ]
    totals = {field: sum(parse_amount(item.get(field)) for item in summaries) for field in numeric_fields}
    totals["storeCount"] = len(summaries)

    top_store = max(summaries, key=lambda item: item["salesTotal"], default=None)
    return jsonify(
        {
            "year": year,
            "stores": summaries,
            "totals": totals,
            "topStore": {
                "name": top_store["storeName"],
                "salesTotal": top_store["salesTotal"],
            }
            if top_store
            else None,
        }
    )


@app.get("/api/reports/store-performance")
def store_performance():
    today_value = date.today()
    year = request.args.get("year", type=int) or today_value.year

    stores = fetch_rows("stores", ["id", "name"], order_by="lower(name), id ASC")
    monthly_totals: list[dict[str, Any]] = []
    for index, month_name in enumerate(MONTH_NAMES, start=1):
        monthly_totals.append(
            {
                "month": index,
                "monthLabel": month_name,
                "salesTotal": 0.0,
                "cashPayments": 0.0,
                "bankPayments": 0.0,
                "expenses": 0.0,
                "salaries": 0.0,
                "totalPayouts": 0.0,
                "operatingDifference": 0.0,
            }
        )

    store_series: list[dict[str, Any]] = []
    for store in stores:
        dataset = build_store_dataset(store["id"], year)
        monthly_sales: list[float] = []
        monthly_cash_payments: list[float] = []
        monthly_operating_difference: list[float] = []
        year_sales_total = 0.0
        year_cash_payments = 0.0
        year_total_payouts = 0.0
        year_operating_difference = 0.0

        for month_index in range(1, 13):
            month_data = filter_dataset_by_month(dataset, month_index)
            summary = build_period_summary(month_data)
            monthly_sales.append(summary["salesTotal"])
            monthly_cash_payments.append(summary["cashPayments"])
            monthly_operating_difference.append(summary["operatingDifference"])

            year_sales_total += summary["salesTotal"]
            year_cash_payments += summary["cashPayments"]
            year_total_payouts += summary["totalPayouts"]
            year_operating_difference += summary["operatingDifference"]

            target = monthly_totals[month_index - 1]
            target["salesTotal"] += summary["salesTotal"]
            target["cashPayments"] += summary["cashPayments"]
            target["bankPayments"] += summary["bankPayments"]
            target["expenses"] += summary["expenses"]
            target["salaries"] += summary["salaries"]
            target["totalPayouts"] += summary["totalPayouts"]
            target["operatingDifference"] += summary["operatingDifference"]

        store_series.append(
            {
                "storeId": store["id"],
                "storeName": store["name"],
                "yearSalesTotal": year_sales_total,
                "yearCashPayments": year_cash_payments,
                "yearTotalPayouts": year_total_payouts,
                "yearOperatingDifference": year_operating_difference,
                "monthlySales": monthly_sales,
                "monthlyCashPayments": monthly_cash_payments,
                "monthlyOperatingDifference": monthly_operating_difference,
            }
        )

    store_series.sort(key=lambda item: item["yearSalesTotal"], reverse=True)
    totals = {
        "salesTotal": sum(item["yearSalesTotal"] for item in store_series),
        "cashPayments": sum(item["yearCashPayments"] for item in store_series),
        "totalPayouts": sum(item["yearTotalPayouts"] for item in store_series),
        "operatingDifference": sum(item["yearOperatingDifference"] for item in store_series),
        "storeCount": len(store_series),
    }

    return jsonify(
        {
            "year": year,
            "months": list(MONTH_NAMES),
            "monthlyTotals": monthly_totals,
            "stores": store_series,
            "totals": totals,
        }
    )


@app.get("/api/lcbo-module/month")
def get_lcbo_month():
    store_id = request.args.get("storeId", type=int)
    if store_id is None:
        abort(400, description="Store is required")
    ensure_store_exists(store_id)
    year, month = validate_year_month(
        request.args.get("year", type=int),
        request.args.get("month", type=int),
    )
    return jsonify(fetch_lcbo_month_payload(store_id, year, month))


@app.post("/api/lcbo-module/validate")
def validate_lcbo_month():
    payload = request.get_json(silent=True) or {}
    store_id = parse_required_int(payload.get("storeId"), "Store")
    ensure_store_exists(store_id)

    year, month = validate_year_month(
        parse_required_int(payload.get("year"), "Year"),
        parse_required_int(payload.get("month"), "Month"),
    )
    period_key = month_period_key(year, month)
    notes = normalize_text(payload.get("notes"))

    db = get_db()
    totals_row = db.execute(
        """
        SELECT COALESCE(SUM(amount + hst), 0) AS lcbo_total
          FROM lcbo_entries
         WHERE store_id = ? AND substr(entry_date, 1, 7) = ?
        """,
        (store_id, period_key),
    ).fetchone()
    default_validated_amount = round_money(totals_row["lcbo_total"])

    if payload.get("validatedAmount") in (None, ""):
        validated_amount = default_validated_amount
    else:
        validated_amount = parse_non_negative_amount(payload.get("validatedAmount"), "Validated amount", required=True)

    with db:
        workflow = ensure_lcbo_workflow_row(db, store_id, year, month)
        if workflow["status"] == LCBO_STATUS_POSTED:
            abort(409, description="This month is already posted. Reverse posting before re-validating.")

        event_type = "REVALIDATED" if workflow["status"] == LCBO_STATUS_VALIDATED else "VALIDATED"
        db.execute(
            """
            UPDATE lcbo_monthly_workflows
               SET status = ?,
                   validated_amount = ?,
                   validated_at = CURRENT_TIMESTAMP,
                   notes = ?,
                   posted_at = NULL,
                   posted_resource = NULL,
                   posted_record_id = NULL,
                   posted_payment_type = NULL,
                   updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (LCBO_STATUS_VALIDATED, round_money(validated_amount), notes, workflow["id"]),
        )
        record_lcbo_workflow_event(
            db,
            workflow["id"],
            event_type,
            workflow["status"],
            LCBO_STATUS_VALIDATED,
            validated_amount,
            notes,
        )

    return jsonify(fetch_lcbo_month_payload(store_id, year, month))


@app.post("/api/lcbo-module/post")
def post_lcbo_month():
    payload = request.get_json(silent=True) or {}
    store_id = parse_required_int(payload.get("storeId"), "Store")
    ensure_store_exists(store_id)

    year, month = validate_year_month(
        parse_required_int(payload.get("year"), "Year"),
        parse_required_int(payload.get("month"), "Month"),
    )
    payment_type = normalize_text(payload.get("paymentType")).upper() or "CASH"
    if payment_type not in {"CASH", "DEBIT"}:
        abort(400, description="Payment type must be Cash or Debit")

    payment_date = normalize_text(payload.get("paymentDate"))
    if payment_date:
        posting_date = validate_iso_date(payment_date, "Payment date")
    else:
        last_day = calendar.monthrange(year, month)[1]
        posting_date = f"{year:04d}-{month:02d}-{last_day:02d}"

    notes = normalize_text(payload.get("notes"))
    posting_label = month_label(year, month)
    posting_description = f"LCBO Payment – {posting_label}"

    db = get_db()
    with db:
        workflow = get_lcbo_workflow_row(db, store_id, year, month)
        if workflow is None or workflow["status"] == LCBO_STATUS_PENDING:
            abort(400, description="Month must be validated before posting")
        if workflow["status"] == LCBO_STATUS_POSTED:
            abort(409, description="This month is already posted to Cash/Debit")

        validated_amount = round_money(workflow["validated_amount"])
        if validated_amount <= 0:
            abort(400, description="Validated amount must be greater than zero before posting")

        source_note = f"LCBO month {month_period_key(year, month)}"
        if notes:
            source_note = f"{source_note} - {notes}"

        if payment_type == "CASH":
            inserted = db.execute(
                """
                INSERT INTO cash_payments (
                    store_id, payment_date, vendor_name, amount, hst, source_type, source_ref_id, source_note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    store_id,
                    posting_date,
                    posting_description,
                    validated_amount,
                    0,
                    LCBO_SOURCE_TYPE,
                    workflow["id"],
                    source_note,
                ),
            )
            posted_resource = "cash-payments"
        else:
            inserted = db.execute(
                """
                INSERT INTO bank_payments (
                    store_id, payment_date, vendor_name, amount, hst, chq, source_type, source_ref_id, source_note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    store_id,
                    posting_date,
                    posting_description,
                    validated_amount,
                    0,
                    "",
                    LCBO_SOURCE_TYPE,
                    workflow["id"],
                    source_note,
                ),
            )
            posted_resource = "bank-payments"

        db.execute(
            """
            UPDATE lcbo_monthly_workflows
               SET status = ?,
                   posted_at = CURRENT_TIMESTAMP,
                   posted_resource = ?,
                   posted_record_id = ?,
                   posted_payment_type = ?,
                   updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (LCBO_STATUS_POSTED, posted_resource, inserted.lastrowid, payment_type, workflow["id"]),
        )
        record_lcbo_workflow_event(
            db,
            workflow["id"],
            "POSTED",
            workflow["status"],
            LCBO_STATUS_POSTED,
            validated_amount,
            notes,
        )

    return jsonify(fetch_lcbo_month_payload(store_id, year, month))


@app.post("/api/lcbo-module/reverse")
def reverse_lcbo_posting():
    payload = request.get_json(silent=True) or {}
    store_id = parse_required_int(payload.get("storeId"), "Store")
    ensure_store_exists(store_id)

    year, month = validate_year_month(
        parse_required_int(payload.get("year"), "Year"),
        parse_required_int(payload.get("month"), "Month"),
    )
    reason = normalize_text(payload.get("reason"))
    db = get_db()
    with db:
        workflow = get_lcbo_workflow_row(db, store_id, year, month)
        if workflow is None or workflow["status"] != LCBO_STATUS_POSTED:
            abort(400, description="Only posted months can be reversed")
        if not workflow["posted_resource"] or not workflow["posted_record_id"]:
            abort(409, description="Linked posted record is missing")

        if workflow["posted_resource"] == "cash-payments":
            deleted = db.execute(
                """
                DELETE FROM cash_payments
                 WHERE id = ? AND source_type = ? AND source_ref_id = ?
                """,
                (workflow["posted_record_id"], LCBO_SOURCE_TYPE, workflow["id"]),
            )
        elif workflow["posted_resource"] == "bank-payments":
            deleted = db.execute(
                """
                DELETE FROM bank_payments
                 WHERE id = ? AND source_type = ? AND source_ref_id = ?
                """,
                (workflow["posted_record_id"], LCBO_SOURCE_TYPE, workflow["id"]),
            )
        else:
            abort(409, description="Unsupported posted resource for reversal")

        if deleted.rowcount == 0:
            abort(409, description="Linked posted record could not be reversed safely")

        db.execute(
            """
            UPDATE lcbo_monthly_workflows
               SET status = ?,
                   posted_at = NULL,
                   posted_resource = NULL,
                   posted_record_id = NULL,
                   posted_payment_type = NULL,
                   updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (LCBO_STATUS_VALIDATED, workflow["id"]),
        )
        record_lcbo_workflow_event(
            db,
            workflow["id"],
            "REVERSED",
            LCBO_STATUS_POSTED,
            LCBO_STATUS_VALIDATED,
            workflow["validated_amount"],
            reason,
        )

    return jsonify(fetch_lcbo_month_payload(store_id, year, month))


@app.get("/api/credit-card-reconciliation")
def list_credit_card_reconciliation():
    board = normalize_text(request.args.get("board")).lower() or "pending"
    store_id = request.args.get("storeId", type=int)
    if board == "store":
        if store_id is None:
            abort(400, description="Store board requires storeId")
        ensure_store_exists(store_id)
    else:
        board = "pending"
        store_id = None

    db = get_db()
    summary_row = db.execute(
        """
        SELECT COALESCE(SUM(CASE WHEN status != ? THEN 1 ELSE 0 END), 0) AS total_open_count,
               COALESCE(SUM(CASE WHEN status != ? AND dedicated_store_id IS NULL THEN amount_cents + hst_cents ELSE 0 END), 0) AS pending_amount_cents,
               COALESCE(SUM(CASE WHEN status = ? THEN amount_cents + hst_cents ELSE 0 END), 0) AS allocated_amount_cents
          FROM credit_card_reconciliation
        """,
        (CCR_STATUS_ALLOCATED, CCR_STATUS_ALLOCATED, CCR_STATUS_ALLOCATED),
    ).fetchone()

    base_select = """
        SELECT c.id, c.transaction_date, c.credit_card, c.merchant_name, c.description, c.amount_cents, c.hst_cents,
               c.dedicated_store_id, c.status, c.allocated_store_id, c.payment_type, c.allocation_resource,
               c.allocation_record_id, c.allocated_at, c.reversed_at, c.created_at, c.updated_at,
               ds.name AS dedicated_store_name, alloc.name AS allocated_store_name
          FROM credit_card_reconciliation c
          LEFT JOIN stores ds ON ds.id = c.dedicated_store_id
          LEFT JOIN stores alloc ON alloc.id = c.allocated_store_id
    """
    params: list[Any] = [CCR_STATUS_ALLOCATED]
    if board == "store":
        where_clause = "WHERE c.status != ? AND c.dedicated_store_id = ?"
        params.append(store_id)
    else:
        where_clause = "WHERE c.status != ? AND c.dedicated_store_id IS NULL"
    rows = db.execute(
        f"""
        {base_select}
        {where_clause}
        ORDER BY c.transaction_date DESC, c.id DESC
        """,
        params,
    ).fetchall()

    stores_rows = db.execute("SELECT id, name FROM stores ORDER BY id ASC LIMIT ?", (STORE_BUTTON_LIMIT,)).fetchall()
    pending_row = db.execute(
        """
        SELECT COUNT(*) AS row_count,
               COALESCE(SUM(amount_cents + hst_cents), 0) AS amount_cents
          FROM credit_card_reconciliation
         WHERE status != ? AND dedicated_store_id IS NULL
        """,
        (CCR_STATUS_ALLOCATED,),
    ).fetchone()
    boards: list[dict[str, Any]] = [
        {
            "type": "pending",
            "label": "PENDING",
            "count": pending_row["row_count"],
            "amount": cents_to_amount(pending_row["amount_cents"]),
            "storeId": None,
        }
    ]
    for store in stores_rows:
        row = db.execute(
            """
            SELECT COUNT(*) AS row_count,
                   COALESCE(SUM(amount_cents + hst_cents), 0) AS amount_cents
              FROM credit_card_reconciliation
             WHERE status != ? AND dedicated_store_id = ?
            """,
            (CCR_STATUS_ALLOCATED, store["id"]),
        ).fetchone()
        boards.append(
            {
                "type": "store",
                "label": store["name"],
                "count": row["row_count"],
                "amount": cents_to_amount(row["amount_cents"]),
                "storeId": store["id"],
            }
        )

    if board == "store":
        selected_title = next((f"{item['label']} DEDICATED TRANSACTIONS" for item in boards if item["type"] == "store" and item["storeId"] == store_id), "DEDICATED TRANSACTIONS")
    else:
        selected_title = "PENDING CREDIT CARD TRANSACTIONS"

    return jsonify(
        {
            "transactions": [serialize_credit_card_reconciliation(row) for row in rows],
            "boards": boards,
            "selectedBoard": {
                "type": board,
                "storeId": store_id,
                "title": selected_title,
            },
            "summary": {
                "totalTransactions": summary_row["total_open_count"],
                "pendingAmount": cents_to_amount(summary_row["pending_amount_cents"]),
                "allocatedAmount": cents_to_amount(summary_row["allocated_amount_cents"]),
            },
        }
    )


@app.get("/api/credit-card-reconciliation/<int:record_id>")
def get_credit_card_reconciliation(record_id: int):
    row = get_credit_card_reconciliation_row(record_id)
    return jsonify(serialize_credit_card_reconciliation(row))


@app.post("/api/credit-card-reconciliation")
def create_credit_card_reconciliation():
    payload = request.get_json(silent=True) or {}
    transaction_date = validate_iso_date(payload.get("transactionDate"), "Transaction date")
    credit_card = normalize_text(payload.get("creditCard"))
    merchant = normalize_text(payload.get("merchant"))
    description = normalize_text(payload.get("description"))
    amount_cents = parse_money_to_cents(payload.get("amount"))
    hst_cents = parse_non_negative_money_to_cents(payload.get("hst"))
    dedicated_store_id = parse_optional_store_id(payload.get("dedicatedStoreId"), "Dedicated store")

    if not credit_card:
        abort(400, description="Credit card is required")
    if not merchant:
        abort(400, description="Merchant is required")

    db = get_db()
    cursor = db.execute(
        """
        INSERT INTO credit_card_reconciliation (
            transaction_date, credit_card, merchant_name, description, amount_cents, hst_cents, dedicated_store_id, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (transaction_date, credit_card, merchant, description, amount_cents, hst_cents, dedicated_store_id, CCR_STATUS_UNALLOCATED),
    )
    db.commit()
    row = get_credit_card_reconciliation_row(cursor.lastrowid)
    return jsonify(serialize_credit_card_reconciliation(row)), 201


@app.patch("/api/credit-card-reconciliation/<int:record_id>")
def update_credit_card_reconciliation(record_id: int):
    current = get_credit_card_reconciliation_row(record_id)
    if current["status"] == CCR_STATUS_ALLOCATED:
        abort(400, description="Allocated transactions cannot be edited. Reverse allocation first.")

    payload = request.get_json(silent=True) or {}
    updates: dict[str, Any] = {}
    if "transactionDate" in payload:
        updates["transaction_date"] = validate_iso_date(payload.get("transactionDate"), "Transaction date")
    if "creditCard" in payload:
        credit_card = normalize_text(payload.get("creditCard"))
        if not credit_card:
            abort(400, description="Credit card is required")
        updates["credit_card"] = credit_card
    if "merchant" in payload:
        merchant = normalize_text(payload.get("merchant"))
        if not merchant:
            abort(400, description="Merchant is required")
        updates["merchant_name"] = merchant
    if "description" in payload:
        updates["description"] = normalize_text(payload.get("description"))
    if "amount" in payload:
        updates["amount_cents"] = parse_money_to_cents(payload.get("amount"))
    if "hst" in payload:
        updates["hst_cents"] = parse_non_negative_money_to_cents(payload.get("hst"))
    if "dedicatedStoreId" in payload:
        updates["dedicated_store_id"] = parse_optional_store_id(payload.get("dedicatedStoreId"), "Dedicated store")

    if not updates:
        abort(400, description="No fields to update")

    assignments = ", ".join(f"{column} = ?" for column in updates)
    params = [*updates.values(), record_id]
    db = get_db()
    db.execute(
        f"""
        UPDATE credit_card_reconciliation
           SET {assignments},
               updated_at = CURRENT_TIMESTAMP
         WHERE id = ?
        """,
        params,
    )
    db.commit()
    row = get_credit_card_reconciliation_row(record_id)
    return jsonify(serialize_credit_card_reconciliation(row))


@app.post("/api/credit-card-reconciliation/<int:record_id>/allocate")
def allocate_credit_card_reconciliation(record_id: int):
    payload = request.get_json(silent=True) or {}
    payment_type = normalize_text(payload.get("paymentType")).upper()
    if payment_type not in {"CASH", "DEBIT"}:
        abort(400, description="Payment type must be Cash or Debit")

    row = get_credit_card_reconciliation_row(record_id)
    if row["status"] == CCR_STATUS_ALLOCATED:
        abort(400, description="This transaction is already allocated")

    destination_store_id = parse_optional_store_id(payload.get("destinationStoreId"), "Destination store")
    if destination_store_id is None:
        destination_store_id = row["dedicated_store_id"]
    if destination_store_id is None:
        abort(400, description="Destination store is required")

    amount = cents_to_amount(row["amount_cents"])
    hst = cents_to_amount(row["hst_cents"])
    vendor_name = f"{row['merchant_name']} [CCR]"
    source_note = f"From CCR ({row['credit_card']})"
    if row["description"]:
        source_note = f"{source_note} - {row['description']}"

    db = get_db()
    with db:
        if payment_type == "CASH":
            linked = db.execute(
                """
                INSERT INTO cash_payments (
                    store_id, payment_date, vendor_name, amount, hst, source_type, source_ref_id, source_note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    destination_store_id,
                    row["transaction_date"],
                    vendor_name,
                    amount,
                    hst,
                    CCR_SOURCE_TYPE,
                    record_id,
                    source_note,
                ),
            )
            allocation_resource = "cash-payments"
        else:
            linked = db.execute(
                """
                INSERT INTO bank_payments (
                    store_id, payment_date, vendor_name, amount, hst, chq, source_type, source_ref_id, source_note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    destination_store_id,
                    row["transaction_date"],
                    vendor_name,
                    amount,
                    hst,
                    "",
                    CCR_SOURCE_TYPE,
                    record_id,
                    source_note,
                ),
            )
            allocation_resource = "bank-payments"

        db.execute(
            """
            UPDATE credit_card_reconciliation
               SET status = ?,
                   allocated_store_id = ?,
                   payment_type = ?,
                   allocation_resource = ?,
                   allocation_record_id = ?,
                   allocated_at = CURRENT_TIMESTAMP,
                   reversed_at = NULL,
                   updated_at = CURRENT_TIMESTAMP
             WHERE id = ? AND status != ?
            """,
            (
                CCR_STATUS_ALLOCATED,
                destination_store_id,
                payment_type,
                allocation_resource,
                linked.lastrowid,
                record_id,
                CCR_STATUS_ALLOCATED,
            ),
        )
    updated = get_credit_card_reconciliation_row(record_id)
    return jsonify(serialize_credit_card_reconciliation(updated))


@app.post("/api/credit-card-reconciliation/<int:record_id>/reverse")
def reverse_credit_card_reconciliation(record_id: int):
    row = get_credit_card_reconciliation_row(record_id)
    if row["status"] != CCR_STATUS_ALLOCATED:
        abort(400, description="Only allocated transactions can be reversed")
    if not row["allocation_resource"] or not row["allocation_record_id"]:
        abort(409, description="Linked allocation record is missing")

    db = get_db()
    with db:
        if row["allocation_resource"] == "cash-payments":
            cursor = db.execute(
                """
                DELETE FROM cash_payments
                 WHERE id = ? AND source_type = ? AND source_ref_id = ?
                """,
                (row["allocation_record_id"], CCR_SOURCE_TYPE, record_id),
            )
        elif row["allocation_resource"] == "bank-payments":
            cursor = db.execute(
                """
                DELETE FROM bank_payments
                 WHERE id = ? AND source_type = ? AND source_ref_id = ?
                """,
                (row["allocation_record_id"], CCR_SOURCE_TYPE, record_id),
            )
        else:
            abort(409, description="Unsupported allocation record type")

        if cursor.rowcount == 0:
            abort(409, description="Linked payment record could not be reversed safely")

        db.execute(
            """
            UPDATE credit_card_reconciliation
               SET status = ?,
                   allocated_store_id = NULL,
                   payment_type = NULL,
                   allocation_resource = NULL,
                   allocation_record_id = NULL,
                   reversed_at = CURRENT_TIMESTAMP,
                   updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (CCR_STATUS_REVERSED, record_id),
        )
    updated = get_credit_card_reconciliation_row(record_id)
    return jsonify(serialize_credit_card_reconciliation(updated))


@app.delete("/api/credit-card-reconciliation/<int:record_id>")
def delete_credit_card_reconciliation(record_id: int):
    row = get_credit_card_reconciliation_row(record_id)
    if row["status"] == CCR_STATUS_ALLOCATED:
        abort(
            400,
            description="This transaction is already allocated. Reverse allocation before deleting.",
        )

    cursor = get_db().execute("DELETE FROM credit_card_reconciliation WHERE id = ?", (record_id,))
    if cursor.rowcount == 0:
        abort(404, description="Credit card reconciliation transaction not found")
    get_db().commit()
    return "", 204


@app.get("/api/stores/<int:store_id>/annual-export")
def export_annual_workbook(store_id: int):
    year = request.args.get("year", type=int)
    if year is None:
        abort(400, description="Year is required")

    dataset = build_store_dataset(store_id, year)
    monthly_summaries: list[dict[str, Any]] = []
    for index, month_name in enumerate(MONTH_NAMES, start=1):
        month_data = filter_dataset_by_month(dataset, index)
        summary = build_export_period_summary(month_data)
        summary["monthLabel"] = month_name
        monthly_summaries.append(summary)

    workbook = Workbook()
    build_summary_sheet(workbook, dataset["store"]["name"], year, monthly_summaries)
    for index, month_name in enumerate(MONTH_NAMES, start=1):
        month_data = filter_dataset_by_month(dataset, index)
        build_month_sheet(workbook, month_name, month_data, monthly_summaries[index - 1])
    q_market_rows = [row for row in dataset["expenses"] if normalize_text(row["vendor_name"]).casefold() == "q-market cig"]
    if q_market_rows:
        build_vendor_expense_sheet(workbook, "Q-Market Cig", q_market_rows)
    build_hst_sheet(workbook, monthly_summaries)
    if dataset["lcbo_entries"]:
        build_resource_sheet(
            workbook,
            "LCBO",
            ["Date", "Vendor", "Invoice No", "Credit Ending", "Amount", "HST", "Total"],
            [
                [
                    excel_date(row["entry_date"]),
                    row["vendor_name"],
                    row["invoice_no"],
                    row["credit_ending"],
                    parse_amount(row["amount"]),
                    parse_amount(row["hst"]),
                    parse_amount(row["amount"]) + parse_amount(row["hst"]),
                ]
                for row in dataset["lcbo_entries"]
            ],
            date_columns={1},
            numeric_columns={5, 6, 7},
        )
    if dataset["credit_card_payments"]:
        build_resource_sheet(
            workbook,
            "Credit Cards",
            ["Date", "Purpose", "Amount"],
            [[excel_date(row["payment_date"]), row["purpose"], parse_amount(row["amount"])] for row in dataset["credit_card_payments"]],
            date_columns={1},
            numeric_columns={3},
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{dataset['store']['name']} {year}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.post("/api/stores/<int:store_id>/annual-import")
def import_annual_workbook(store_id: int):
    ensure_store_exists(store_id)
    year = request.args.get("year", type=int) or request.form.get("year", type=int)
    if year is None:
        abort(400, description="Year is required")

    file = request.files.get("file")
    if file is None or normalize_text(file.filename) == "":
        abort(400, description="Excel workbook file is required")

    try:
        workbook = load_workbook(file.stream, data_only=True)
    except (InvalidFileException, BadZipFile, KeyError, ValueError) as error:
        abort(400, description=f"Invalid Excel workbook: {error}")

    parsed = parse_annual_import_workbook(workbook, year)
    summary_counts = {
        "daily_sales": 0,
        "lottery_records": 0,
        "cash_payments": 0,
        "bank_payments": 0,
        "expenses": 0,
        "salaries": 0,
        "other_income": 0,
        "lcbo_entries": None,
        "credit_card_payments": None,
    }

    db = get_db()
    with db:
        for month_index, month_rows in parsed["months"].items():
            daily_rows = [
                (
                    store_id,
                    row["entry_date"],
                    row["total"],
                    row["sales"],
                    row["hst"],
                    row["online"],
                    row["instant"],
                    row["cc"],
                    row["gc"],
                    row["non_add"],
                    row["mc"],
                    row["visa"],
                    row["debit"],
                    row["cash"],
                    row["lottery_payment"],
                    row["lottery_income"],
                )
                for row in month_rows["daily_sales"]
            ]
            replace_month_rows(
                db,
                table="daily_sales",
                date_col="entry_date",
                store_id=store_id,
                year=year,
                month=month_index,
                columns=[
                    "store_id",
                    "entry_date",
                    "total",
                    "sales",
                    "hst",
                    "online",
                    "instant",
                    "cc",
                    "gc",
                    "non_add",
                    "mc",
                    "visa",
                    "debit",
                    "cash",
                    "lottery_payment",
                    "lottery_income",
                ],
                rows=daily_rows,
            )
            summary_counts["daily_sales"] += len(daily_rows)

            lottery_rows = [
                (store_id, row["lottery_date"], row["lottery_payment"], row["lottery_income"])
                for row in month_rows["lottery_records"]
            ]
            replace_month_rows(
                db,
                table="lottery_records",
                date_col="lottery_date",
                store_id=store_id,
                year=year,
                month=month_index,
                columns=["store_id", "lottery_date", "lottery_payment", "lottery_income"],
                rows=lottery_rows,
            )
            summary_counts["lottery_records"] += len(lottery_rows)

            cash_rows = [
                (
                    store_id,
                    row["payment_date"],
                    row["vendor_name"],
                    row["amount"],
                    row["hst"],
                    "excel_import",
                    None,
                    f"Annual workbook import {year}",
                )
                for row in month_rows["cash_payments"]
            ]
            replace_month_rows(
                db,
                table="cash_payments",
                date_col="payment_date",
                store_id=store_id,
                year=year,
                month=month_index,
                columns=[
                    "store_id",
                    "payment_date",
                    "vendor_name",
                    "amount",
                    "hst",
                    "source_type",
                    "source_ref_id",
                    "source_note",
                ],
                rows=cash_rows,
            )
            summary_counts["cash_payments"] += len(cash_rows)

            bank_rows = [
                (
                    store_id,
                    row["payment_date"],
                    row["vendor_name"],
                    row["amount"],
                    row["hst"],
                    row["chq"],
                    "excel_import",
                    None,
                    f"Annual workbook import {year}",
                )
                for row in month_rows["bank_payments"]
            ]
            replace_month_rows(
                db,
                table="bank_payments",
                date_col="payment_date",
                store_id=store_id,
                year=year,
                month=month_index,
                columns=[
                    "store_id",
                    "payment_date",
                    "vendor_name",
                    "amount",
                    "hst",
                    "chq",
                    "source_type",
                    "source_ref_id",
                    "source_note",
                ],
                rows=bank_rows,
            )
            summary_counts["bank_payments"] += len(bank_rows)

            expense_rows = [
                (store_id, row["expense_date"], row["vendor_name"], row["amount"], row["hst"])
                for row in month_rows["expenses"]
            ]
            replace_month_rows(
                db,
                table="expenses",
                date_col="expense_date",
                store_id=store_id,
                year=year,
                month=month_index,
                columns=["store_id", "expense_date", "vendor_name", "amount", "hst"],
                rows=expense_rows,
            )
            summary_counts["expenses"] += len(expense_rows)

            salary_rows = [
                (store_id, row["salary_date"], row["employee"], row["amount"])
                for row in month_rows["salaries"]
            ]
            replace_month_rows(
                db,
                table="salaries",
                date_col="salary_date",
                store_id=store_id,
                year=year,
                month=month_index,
                columns=["store_id", "salary_date", "employee", "amount"],
                rows=salary_rows,
            )
            summary_counts["salaries"] += len(salary_rows)

            other_income_rows = [
                (store_id, row["income_date"], row["vendor_name"], row["amount"])
                for row in month_rows["other_income"]
            ]
            replace_month_rows(
                db,
                table="other_income",
                date_col="income_date",
                store_id=store_id,
                year=year,
                month=month_index,
                columns=["store_id", "income_date", "vendor_name", "amount"],
                rows=other_income_rows,
            )
            summary_counts["other_income"] += len(other_income_rows)

        if parsed["lcbo_entries"] is not None:
            lcbo_rows = [
                (
                    store_id,
                    row["entry_date"],
                    "LCBO",
                    row["invoice_no"],
                    row["credit_ending"],
                    row["amount"],
                    row["hst"],
                )
                for row in parsed["lcbo_entries"]
            ]
            replace_year_rows(
                db,
                table="lcbo_entries",
                date_col="entry_date",
                store_id=store_id,
                year=year,
                columns=["store_id", "entry_date", "vendor_name", "invoice_no", "credit_ending", "amount", "hst"],
                rows=lcbo_rows,
            )
            summary_counts["lcbo_entries"] = len(lcbo_rows)

        if parsed["credit_card_payments"] is not None:
            cc_rows = [
                (store_id, row["payment_date"], row["purpose"], row["amount"])
                for row in parsed["credit_card_payments"]
            ]
            replace_year_rows(
                db,
                table="credit_card_payments",
                date_col="payment_date",
                store_id=store_id,
                year=year,
                columns=["store_id", "payment_date", "purpose", "amount"],
                rows=cc_rows,
            )
            summary_counts["credit_card_payments"] = len(cc_rows)

    return jsonify(
        {
            "status": "ok",
            "storeId": store_id,
            "year": year,
            "imported": summary_counts,
            "monthsImported": sorted(parsed["months"].keys()),
        }
    )


@app.get("/api/stores/<int:store_id>/<resource>")
def list_resource(store_id: int, resource: str):
    ensure_store_exists(store_id)
    config = get_resource_config(resource)
    where_sql, params = build_date_filters(store_id, config["date_col"])
    rows = get_db().execute(
        f"""
        SELECT id, {", ".join(config["mapping"].values())}
          FROM {config["table"]}
         WHERE {where_sql}
         ORDER BY {config["date_col"]} DESC, id DESC
        """,
        params,
    ).fetchall()
    return jsonify([serialize_resource(resource, row) for row in rows])


@app.post("/api/stores/<int:store_id>/<resource>")
def create_resource(store_id: int, resource: str):
    ensure_store_exists(store_id)
    config = get_resource_config(resource)
    payload = request.get_json(silent=True) or {}
    require_fields(payload, config["required"])
    values = map_payload(payload, config["mapping"], config["numeric"], partial=False)
    if resource == "lcbo-entries":
        values["vendor_name"] = "LCBO"

    columns = ["store_id", *values.keys()]
    placeholders = ", ".join("?" for _ in columns)

    db = get_db()
    if resource in LCBO_LOCKED_RESOURCES:
        enforce_lcbo_month_editable(
            db,
            store_id,
            values[config["date_col"]],
            f"{resource} record added",
        )
    try:
        cursor = db.execute(
            f"INSERT INTO {config['table']} ({', '.join(columns)}) VALUES ({placeholders})",
            (store_id, *values.values()),
        )
    except sqlite3.IntegrityError as error:
        abort(400, description=str(error))
    db.commit()

    row = db.execute(
        f"SELECT id, {', '.join(config['mapping'].values())} FROM {config['table']} WHERE id = ?",
        (cursor.lastrowid,),
    ).fetchone()
    return jsonify(serialize_resource(resource, row)), 201


@app.patch("/api/<resource>/<int:record_id>")
def update_resource(resource: str, record_id: int):
    config = get_resource_config(resource)
    payload = request.get_json(silent=True) or {}
    values = map_payload(payload, config["mapping"], config["numeric"], partial=True)
    if resource == "lcbo-entries" and "vendorName" in payload:
        values["vendor_name"] = "LCBO"
    if not values:
        abort(400, description="No fields to update")

    assignments = ", ".join(f"{column} = ?" for column in values)
    db = get_db()
    existing = db.execute(
        f"SELECT id, store_id, {config['date_col']} AS date_value FROM {config['table']} WHERE id = ?",
        (record_id,),
    ).fetchone()
    if existing is None:
        abort(404, description="Record not found")

    if resource in LCBO_LOCKED_RESOURCES:
        enforce_lcbo_month_editable(
            db,
            existing["store_id"],
            existing["date_value"],
            f"{resource} record updated",
        )
        new_date_value = values.get(config["date_col"])
        if new_date_value and new_date_value != existing["date_value"]:
            enforce_lcbo_month_editable(
                db,
                existing["store_id"],
                new_date_value,
                f"{resource} record moved to another month",
            )
    try:
        cursor = db.execute(
            f"""
            UPDATE {config['table']}
               SET {assignments},
                   updated_at = CURRENT_TIMESTAMP
             WHERE id = ?
            """,
            (*values.values(), record_id),
        )
    except sqlite3.IntegrityError as error:
        abort(400, description=str(error))
    if cursor.rowcount == 0:
        abort(404, description="Record not found")

    db.commit()
    row = db.execute(
        f"SELECT id, {', '.join(config['mapping'].values())} FROM {config['table']} WHERE id = ?",
        (record_id,),
    ).fetchone()
    return jsonify(serialize_resource(resource, row))


@app.delete("/api/<resource>/<int:record_id>")
def delete_resource(resource: str, record_id: int):
    config = get_resource_config(resource)
    db = get_db()
    existing = db.execute(
        f"SELECT id, store_id, {config['date_col']} AS date_value FROM {config['table']} WHERE id = ?",
        (record_id,),
    ).fetchone()
    if existing is None:
        abort(404, description="Record not found")
    if resource in LCBO_LOCKED_RESOURCES:
        enforce_lcbo_month_editable(
            db,
            existing["store_id"],
            existing["date_value"],
            f"{resource} record deleted",
        )

    cursor = db.execute(f"DELETE FROM {config['table']} WHERE id = ?", (record_id,))
    if cursor.rowcount == 0:
        abort(404, description="Record not found")
    db.commit()
    return "", 204


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run Sales Reporting Dashboard")
    parser.add_argument("--host", default=os.environ.get("HOST", "0.0.0.0"))
    parser.add_argument("--port", type=int, default=int(os.environ.get("PORT", "8080")))
    args = parser.parse_args()

    ensure_app_initialized()
    app.run(host=args.host, port=args.port, debug=False)
